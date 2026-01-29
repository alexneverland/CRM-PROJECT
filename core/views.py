# core/views.py
# --- 1. Python & Django Standard Libraries ---
import datetime
import io
import unicodedata
from collections import defaultdict
from decimal import Decimal
from datetime import timedelta
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model, logout as auth_logout_function
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.core.mail import EmailMessage
from django.forms import inlineformset_factory
from django.http import JsonResponse, HttpResponse, Http404
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET
from django.http import Http404, JsonResponse
import json
from django.db.models.functions import TruncMonth
from django.db.models import Sum, F, Q, DecimalField, ExpressionWrapper, Case, When, Value, Count
from django.db.models.functions import Coalesce
from .models import Attachment 
from .forms import AttachmentForm 
from django.contrib.contenttypes.models import ContentType

# --- 2. External Libraries ---
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl.utils import get_column_letter

# --- 3. Local Application Imports (From this project) ---
from .models import (
    Customer, Product, Order, OrderItem, StockReceipt, Payment, ActivityLog, 
    SalesRepresentative, Invoice, InvoiceItem, Commission, UserProfile,
    CreditNote, CreditNoteItem, RetailReceipt, RetailReceiptItem, DeliveryNote, DeliveryNoteItem, Supplier, PurchaseOrder
)
from .forms import (
    CustomerForm, ProductForm, OrderForm, OrderItemForm, 
    StockReceiptForm, PaymentForm, PaymentCancellationForm,
    CustomUserCreationForm, CustomUserChangeForm, MyProfileForm,
    UserProfileForm, CreditNoteForm, CreditNoteItemFormSet,
    RetailReceiptItemFormSet, DeliveryNoteEditForm, StandaloneDeliveryNoteForm, StandaloneDeliveryNoteItemFormSet, InvoiceEditForm, RetailReceiptForm,
    SalesRepresentativeForm, SupplierForm, PurchaseOrderForm, PurchaseOrderItemFormSet, PurchaseOrder, PurchaseOrderItem, ReceivePOItemFormSet
)


# --- 4. WeasyPrint (Optional PDF Library) ---
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False
    import logging
    logger = logging.getLogger(__name__)
    logger.warning("WeasyPrint library is not installed. PDF generation will not be available.")


# --- Global Variables & Helper Functions ---
User = get_user_model()

def normalize_text(text):
    if not text:
        return ""
    text = unicodedata.normalize('NFD', text)
    return ''.join(c for c in text if unicodedata.category(c) != 'Mn').lower()

# (Η inlineformset_factory μεταφέρθηκε εδώ από το παλιό σου αρχείο, για συνέπεια)
OrderItemFormSet = inlineformset_factory(
    Order, OrderItem,
    form=OrderItemForm,
    fields=('product', 'quantity', 'is_gift', 'unit_price', 'discount_percentage', 'vat_percentage', 'comments'),
    extra=1,
    can_delete=True
)

# 
# --- VIEWS ΞΕΚΙΝΟΥΝ ΑΠΟ ΕΔΩ ---
#
@login_required
def home(request):
    low_stock_products = Product.objects.filter(stock_quantity__lte=F('min_stock_level')).order_by('stock_quantity')
    low_stock_count = low_stock_products.count()
    total_customers = Customer.objects.count()
    total_products = Product.objects.count()
    open_orders_count = Order.objects.filter(status__in=[Order.STATUS_PENDING, Order.STATUS_PROCESSING]).count()
    today = timezone.now().date()
    sales_this_month = Order.objects.filter(status=Order.STATUS_COMPLETED, order_date__year=today.year, order_date__month=today.month).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0.00
    recent_orders = Order.objects.select_related('customer').order_by('-order_date', '-pk')[:5]
    recent_payments = Payment.objects.filter(status=Payment.STATUS_ACTIVE).select_related('customer').order_by('-payment_date', '-pk')[:5]

    # --- ΝΕΑ ΛΟΓΙΚΗ ΓΙΑ ΤΙΜΟΛΟΓΙΑ ΠΟΥ ΛΗΓΟΥΝ ΣΥΝΤΟΜΑ ---
    due_date_threshold = today + timedelta(days=7) # Ορίζουμε το "σύντομα" ως 7 ημέρες
    due_soon_invoices = Invoice.objects.filter(
        status=Invoice.STATUS_ISSUED, # Θέλουμε μόνο τα ανεξόφλητα
        due_date__gte=today,          # που δεν έχουν λήξει ακόμα
        due_date__lte=due_date_threshold # αλλά λήγουν μέχρι το όριο που θέσαμε
    ).select_related('customer').order_by('due_date')
    due_soon_count = due_soon_invoices.count()
    # --- ΤΕΛΟΣ ΝΕΑΣ ΛΟΓΙΚΗΣ ---

    context = {
        'title': 'Dashboard - Αρχική',
        'low_stock_products': low_stock_products,
        'low_stock_count': low_stock_count,
        'total_customers': total_customers,
        'total_products': total_products,
        'open_orders_count': open_orders_count,
        'sales_this_month': sales_this_month,
        'recent_orders': recent_orders,
        'recent_payments': recent_payments,
        'due_soon_invoices': due_soon_invoices, # Προσθήκη στο context
        'due_soon_count': due_soon_count,     # Προσθήκη στο context
    }
    return render(request, 'core/home.html', context)

@login_required
def customer_create(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save()
            messages.success(request, f'Ο πελάτης "{customer}" προστέθηκε επιτυχώς.')
            return redirect('customer_detail', pk=customer.pk)
    else:
        form = CustomerForm()
    return render(request, 'core/customer_create.html', {'form': form})
@login_required
def customer_list(request):
    user = request.user
    
    # Ξεκινάμε με ένα queryset που θα περιέχει όλους τους πελάτες
    customers_qs = Customer.objects.all()

    # Εφαρμόζουμε τους κανόνες δικαιωμάτων
    # Αν ο χρήστης ΔΕΝ είναι superuser, τότε εφαρμόζουμε περιορισμούς
    if not user.is_superuser:
        # Αν ανήκει στην ομάδα των πωλητών
        if user.groups.filter(name='Πωλητές').exists():
            try:
                # Φιλτράρουμε τους πελάτες ώστε να δει μόνο αυτούς που του ανήκουν
                customers_qs = customers_qs.filter(sales_rep=user.salesrepresentative)
            except SalesRepresentative.DoesNotExist:
                # Αν για κάποιο λόγο είναι στην ομάδα αλλά δεν έχει προφίλ πωλητή, δεν βλέπει κανέναν
                customers_qs = Customer.objects.none()
        
        # Αν ο χρήστης ανήκει στην ομάδα της Αποθήκης (μπορεί να δει πελάτες αλλά όχι οικονομικά)
        # elif user.groups.filter(name='Αποθήκη').exists():
        #     customers_qs = Customer.objects.all() # Του επιτρέπουμε να τους δει όλους

        # Αν δεν ανήκει σε καμία ειδική ομάδα (π.χ. απλός χρήστης χωρίς δικαιώματα), δεν βλέπει τίποτα
        # else:
        #     customers_qs = Customer.objects.none()

    # Από εδώ και κάτω, η λογική της αναζήτησης και σελιδοποίησης
    # εφαρμόζεται πάνω στο queryset που έχουμε ήδη φιλτράρει (customers_qs)
    query_from_request = request.GET.get('q', None)
    if query_from_request:
        query = query_from_request.strip()
        if query:
            normalized_user_query = normalize_text(query)
            customers_qs = customers_qs.filter(
                Q(first_name_normalized__icontains=normalized_user_query) |
                Q(last_name_normalized__icontains=normalized_user_query) |
                Q(company_name_normalized__icontains=normalized_user_query) |
                Q(code__icontains=query) |
                Q(email__icontains=query) |
                Q(phone__icontains=query) |
                # --- ΝΕΑ ΠΡΟΣΘΗΚΗ: ΨΑΞΕ ΚΑΙ ΣΤΟ ΟΝΟΜΑ ΤΟΥ ΚΕΝΤΡΙΚΟΥ ---
                Q(parent__company_name_normalized__icontains=normalized_user_query) |
                Q(parent__first_name_normalized__icontains=normalized_user_query) |
                Q(parent__last_name_normalized__icontains=normalized_user_query)
            ).distinct()

    paginator = Paginator(customers_qs.order_by('company_name', 'last_name'), 20)
    page_number = request.GET.get('page')
    customers_page = paginator.get_page(page_number)
    
    context = {
        'customers': customers_page,
        'query': query_from_request,
        'title': "Λίστα Πελατών"
    }
    return render(request, 'core/customers_list.html', context)

@login_required
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            customer = form.save()
            messages.success(request, f'Οι αλλαγές στον πελάτη "{customer}" αποθηκεύτηκαν επιτυχώς.')
            return redirect('customer_detail', pk=customer.pk)
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'core/customer_edit.html', {'form': form, 'customer': customer})

@login_required
@require_POST
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    customer_name = str(customer)
    customer.delete()
    messages.success(request, f'Ο πελάτης "{customer_name}" διαγράφηκε επιτυχώς.')
    return redirect('customer_list')

# --- VIEWS ΠΡΟΪΟΝΤΩΝ ---
@login_required
def product_list(request):
    query = request.GET.get('q')
    
    # Το αρχικό queryset παραμένει το ίδιο
    products_qs = Product.objects.all().order_by('name')

    if query:
        query = query.strip()
        # Η συνάρτηση normalize_text πρέπει να είναι διαθέσιμη εδώ
        normalized_query = normalize_text(query)
        products_qs = products_qs.filter(
            Q(name_normalized__icontains=normalized_query) |
            Q(code__icontains=query) |
            Q(barcode__icontains=query) |
            Q(description__icontains=query) 
        ).distinct()

    # --- ΝΕΑ ΛΟΓΙΚΗ ΓΙΑ ΣΕΛΙΔΟΠΟΙΗΣΗ ---
    paginator = Paginator(products_qs, 20) # Π.χ., 20 προϊόντα ανά σελίδα
    page_number = request.GET.get('page')
    try:
        products_page = paginator.page(page_number)
    except PageNotAnInteger:
        products_page = paginator.page(1)
    except EmptyPage:
        products_page = paginator.page(paginator.num_pages)
    
    # --- ΤΕΛΟΣ ΝΕΑΣ ΛΟΓΙΚΗΣ ---

    context = {
        'products': products_page, # <<< ΑΛΛΑΓΗ: Περνάμε το page object
        'query': query,
        'title': "Λίστα Προϊόντων"
    }
    return render(request, 'core/products_list.html', context)

@login_required
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    
    # Ανάκτηση των OrderItem για αυτό το προϊόν από ΟΛΟΚΛΗΡΩΜΕΝΕΣ παραγγελίες
    # Ταξινομημένα από την πιο πρόσφατη παραγγελία
    sales_history_items = OrderItem.objects.filter(
        product=product,
        order__status=Order.STATUS_COMPLETED  # Φιλτράρουμε βάσει της κατάστασης της παραγγελίας
    ).select_related('order', 'order__customer').order_by('-order__order_date', '-order__pk')

    context = {
        'product': product,
        'sales_history_items': sales_history_items, # Προσθήκη του ιστορικού πωλήσεων στο context
        'title': f'Προϊόν: {product.name}'
    }
    return render(request, 'core/product_detail.html', context)

@login_required
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save()
            messages.success(request, f'Το προϊόν "{product}" προστέθηκε επιτυχώς.')
            return redirect('product_detail', pk=product.pk)
    else:
        form = ProductForm()
    return render(request, 'core/product_create.html', {'form': form})

@login_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            product = form.save()
            messages.success(request, f'Οι αλλαγές στο προϊόν "{product}" αποθηκεύτηκαν επιτυχώς.')
            return redirect('product_detail', pk=product.pk)
    else:
        form = ProductForm(instance=product)
    
    context = {
        'form': form,
        'product': product,
        'title': f'Επεξεργασία Προϊόντος: {product.name}'
    }
    return render(request, 'core/product_edit.html', context)

@login_required
@require_POST
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    product_name = str(product)
    product.delete()
    messages.success(request, f'Το προϊόν "{product_name}" διαγράφηκε επιτυχώς.')
    return redirect('product_list')




@login_required
def order_list(request):
    # Η αρχή παραμένει ίδια
    user = request.user
    orders_qs = Order.objects.select_related('customer').all()

    if not user.is_superuser:
        if user.groups.filter(name='Πωλητές').exists():
            try:
                orders_qs = orders_qs.filter(customer__sales_rep=user.salesrepresentative)
            except SalesRepresentative.DoesNotExist:
                orders_qs = Order.objects.none()

    # Τα φίλτρα παραμένουν ίδια
    status_filter = request.GET.get('status', 'all')
    query_text = request.GET.get('q', '').strip()
    customer_filter_id = request.GET.get('customer_filter', None)
    date_from_str = request.GET.get('date_from', None)
    date_to_str = request.GET.get('date_to', None)
    invoiced_filter = request.GET.get('invoiced')
    delivery_note_filter = request.GET.get('delivery_note')

    # Η λογική φιλτραρίσματος παραμένει ίδια
    if status_filter != 'all' and status_filter:
        orders_qs = orders_qs.filter(status=status_filter)
    if customer_filter_id:
        try:
            orders_qs = orders_qs.filter(customer_id=int(customer_filter_id))
        except ValueError:
            customer_filter_id = None
    if date_from_str:
        try:
            orders_qs = orders_qs.filter(order_date__gte=date_from_str)
        except ValueError:
            date_from_str = None 
    if date_to_str:
        try:
            orders_qs = orders_qs.filter(order_date__lte=date_to_str)
        except ValueError:
            date_to_str = None
    if query_text:
        normalized_query = normalize_text(query_text)
        orders_qs = orders_qs.filter(
            Q(order_number__icontains=query_text) |
            Q(customer__first_name_normalized__icontains=normalized_query) |
            Q(customer__last_name_normalized__icontains=normalized_query) |
            Q(customer__company_name_normalized__icontains=normalized_query) |
            Q(customer__code__icontains=query_text)
        ).distinct()

    if invoiced_filter == 'yes':
        orders_qs = orders_qs.filter(invoice__isnull=False)
    elif invoiced_filter == 'no':
        orders_qs = orders_qs.filter(invoice__isnull=True)

    if delivery_note_filter == 'yes':
        orders_qs = orders_qs.filter(delivery_notes__isnull=False)
    elif delivery_note_filter == 'no':
        orders_qs = orders_qs.filter(delivery_notes__isnull=True)

    # --- Η ΓΡΑΜΜΗ ΠΟΥ ΕΛΕΙΠΕ ---
    customers_for_filter = Customer.objects.all().order_by('company_name', 'last_name', 'first_name')
    
    paginator = Paginator(orders_qs.order_by('-order_date', '-pk'), 15)
    page_number = request.GET.get('page')
    orders_page = paginator.get_page(page_number)

    context = {
        'orders': orders_page,
        'status_filter': status_filter,
        'query': query_text,
        'customer_filter_id': customer_filter_id,
        'date_from_value': date_from_str,
        'date_to_value': date_to_str,
        'ORDER_STATUS_CHOICES': Order.ORDER_STATUS_CHOICES,
        'title': 'Λίστα Παραγγελιών',
        # --- Η ΠΡΟΣΘΗΚΗ ΠΟΥ ΕΛΕΙΠΕ ΣΤΟ CONTEXT ---
        'customers_for_filter': customers_for_filter,
        'invoiced_filter': invoiced_filter,
        'delivery_note_filter': delivery_note_filter,
    }
    return render(request, 'core/order_list.html', context)

@login_required
def order_create(request):
    # --- ΝΕΑ ΛΟΓΙΚΗ ΓΙΑ ΠΡΟΣΥΜΠΛΗΡΩΣΗ ΠΕΛΑΤΗ ---
    customer_id = request.GET.get('customer_id')
    customer = None
    initial_data = {'order_date': timezone.now().date()}
    if customer_id:
        try:
            customer = get_object_or_404(Customer, pk=customer_id)
            initial_data['customer'] = customer
            initial_data['shipping_name'] = customer.shipping_address or customer.get_full_name()
            initial_data['shipping_address'] = customer.shipping_address or customer.address
            initial_data['shipping_city'] = customer.shipping_city or customer.city
            initial_data['shipping_postal_code'] = customer.shipping_postal_code or customer.postal_code
        except Http404:
            messages.error(request, "Ο επιλεγμένος πελάτης δεν βρέθηκε.")
    
    if request.method == 'POST':
        form = OrderForm(request.POST)
        formset = OrderItemFormSet(request.POST, prefix='items')
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                # Η λογική αποθήκευσης παραμένει ως έχει
                order = form.save()
                formset.instance = order
                formset.save()
            messages.success(request, f'Η παραγγελία {order.order_number} δημιουργήθηκε επιτυχώς.')
            return redirect('order_detail', pk=order.pk)
        else:
            messages.error(request, "Η φόρμα περιέχει σφάλματα. Παρακαλώ διορθώστε τα.")
    else:
        # Χρησιμοποιούμε τα initial_data που φτιάξαμε παραπάνω
        form = OrderForm(initial=initial_data)
        formset = OrderItemFormSet(prefix='items', queryset=OrderItem.objects.none())

    context = {
        'form': form,
        'formset': formset,
        'title': 'Δημιουργία Νέας Παραγγελίας',
        # Περνάμε και το customer object στο template για να εμφανίσουμε το όνομά του
        'customer': customer, 
    }
    return render(request, 'core/order_create.html', context)


@login_required
def order_edit(request, pk):
    order = get_object_or_404(Order, pk=pk)

    # --- ΟΡΙΣΤΙΚΟΣ ΕΛΕΓΧΟΣ ΚΛΕΙΔΩΜΑΤΟΣ ΠΑΡΑΓΓΕΛΙΑΣ ---
    # Παίρνουμε το τελευταίο Δ.Α. που συνδέεται με την παραγγελία.
    last_dn = order.delivery_notes.last()
    # Η παραγγελία κλειδώνει αν έχει τιμολόγιο, ή αν έχει Δ.Α. το οποίο ΔΕΝ είναι ακυρωμένο.
    has_invoice = hasattr(order, 'invoice') and order.invoice is not None
    has_active_delivery_note = last_dn and last_dn.status != 'CANCELLED'

    if has_invoice or has_active_delivery_note:
        error_message = "Η παραγγελία δεν μπορεί να επεξεργαστεί"
        if has_invoice:
            error_message += f" γιατί έχει ήδη εκδοθεί το τιμολόγιο {order.invoice.invoice_number}."
        elif has_active_delivery_note:
            error_message += f" γιατί έχει ήδη εκδοθεί το Δελτίο Αποστολής {last_dn.delivery_note_number}."
        
        messages.error(request, error_message)
        return redirect('order_detail', pk=order.pk)
    # --- ΤΕΛΟΣ ΕΛΕΓΧΟΥ ---

    if request.method == 'POST':
        form = OrderForm(request.POST, instance=order)
        formset = OrderItemFormSet(request.POST, instance=order, prefix='items')
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                form.save()
                formset.save()
            messages.success(request, f'Οι αλλαγές στην παραγγελία {order.order_number} αποθηκεύτηκαν επιτυχώς.')
            return redirect('order_detail', pk=order.pk)
        else:
            messages.error(request, "Η φόρμα περιέχει σφάλματα. Παρακαλώ διορθώστε τα.")
    else:
        form = OrderForm(instance=order)
        formset = OrderItemFormSet(instance=order, prefix='items')

    context = {
        'form': form,
        'formset': formset,
        'order': order,
        'title': f'Επεξεργασία Παραγγελίας: {order.order_number}'
    }
    return render(request, 'core/order_edit.html', context)

@login_required
def order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)
    
    # Υπάρχουσα λογική (σωστή)
    has_active_invoice = hasattr(order, 'invoice') and order.invoice and order.invoice.status != 'CANCELLED'
    has_active_delivery_note = order.delivery_notes.filter(status__in=['PREPARING', 'SHIPPED', 'DELIVERED']).exists()
    
    is_order_editable = not (has_active_invoice or has_active_delivery_note)

    # --- ΟΡΙΣΤΙΚΗ ΛΟΓΙΚΗ ΓΙΑ ΤΟ ΚΟΥΜΠΙ Δ.Α. ---
    # Μπορούμε να φτιάξουμε Δ.Α. ΜΟΝΟ αν:
    # 1. Η παραγγελία είναι 'Ολοκληρωμένη' ΚΑΙ
    # 2. Δεν υπάρχει ήδη ενεργό Δ.Α.
    can_create_dn = (order.status == Order.STATUS_COMPLETED) and (not has_active_delivery_note)

    context = {
        'order': order, 
        'order_items': order.items.all(),
        'is_order_editable': is_order_editable, 
        'can_create_delivery_note': can_create_dn # Χρησιμοποιούμε τη νέα, σωστή μεταβλητή
    }
     # --- ΠΡΟΣΘΕΣΕ ΑΥΤΟ ΤΟ ΚΟΜΜΑΤΙ ΓΙΑ DEBUG ---
    print("--- DEBUG VIEW ---")
    print(f"Order ID: {order.pk}")
    print(f"Order Status: '{order.status}'")
    print(f"has_active_delivery_note: {has_active_delivery_note}")
    print(f"CAN CREATE DN (Final Decision in View): {can_create_dn}")
    print("------------------")
    # --- ΤΕΛΟΣ DEBUG ---
    return render(request, 'core/order_detail.html', context)


@login_required
@require_POST
def order_delete(request, pk):
    order = get_object_or_404(Order, pk=pk)
    order_number = order.order_number; order.delete()
    messages.success(request, f'Η παραγγελία {order_number} διαγράφηκε επιτυχώς.')
    return redirect('order_list')


# --- AJAX VIEWS ---
@require_GET
@login_required
def get_customer_details_ajax(request, pk):
    try:
        customer = Customer.objects.select_related('parent').get(pk=pk)
        
        # --- ΕΔΩ ΕΙΝΑΙ Η ΔΕΥΤΕΡΗ ΔΙΟΡΘΩΣΗ ---
        data = {
            'pk': customer.pk,
            'name': str(customer),
            'is_branch': customer.is_branch,
            
            # Χρησιμοποιούμε τα βασικά πεδία address, city, postal_code
            'shipping_name': customer.company_name or customer.get_full_name(),
            'shipping_address': customer.address, # <-- ΑΛΛΑΓΗ
            'shipping_city': customer.city, # <-- ΑΛΛΑΓΗ
            'shipping_postal_code': customer.postal_code, # <-- ΑΛΛΑΓΗ
            
            'parent': None
        }

        if customer.is_branch and customer.parent:
            parent = customer.parent
            data['parent'] = {
                'name': str(parent),
                'address': parent.address,
                'city': parent.city,
                'vat_number': parent.vat_number
            }
            
        return JsonResponse(data)
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Customer not found'}, status=404)
@require_GET
@login_required
def search_customers_ajax(request):
    query_from_request = request.GET.get('q', None)
    initial_id = request.GET.get('initial_id', None)
    customers_data = []
    
    customers_found_qs = Customer.objects.all()

    if initial_id:
        customers_found_qs = customers_found_qs.filter(pk=int(initial_id))
    elif query_from_request:
        query = query_from_request.strip()
        if query:
            normalized_user_query = normalize_text(query)
            customers_found_qs = customers_found_qs.filter(
                Q(first_name_normalized__icontains=normalized_user_query) |
                Q(last_name_normalized__icontains=normalized_user_query) |
                Q(company_name_normalized__icontains=normalized_user_query) |
                Q(code__icontains=query) |
                Q(parent__company_name_normalized__icontains=normalized_user_query) |
                Q(parent__first_name_normalized__icontains=normalized_user_query) |
                Q(parent__last_name_normalized__icontains=normalized_user_query)
            ).distinct()
    
    customers_found = customers_found_qs.order_by('company_name', 'last_name', 'first_name')[:15]

    for customer_obj in customers_found:
        # --- ΕΔΩ ΕΙΝΑΙ Η ΔΙΟΡΘΩΣΗ ---
        # Χρησιμοποιούμε τα βασικά πεδία address, city, postal_code
        customers_data.append({
            'id': customer_obj.pk,
            'text': str(customer_obj), 
            'shipping_name': customer_obj.company_name or customer_obj.get_full_name(),
            'shipping_address': customer_obj.address, # <-- ΑΛΛΑΓΗ
            'shipping_city': customer_obj.city, # <-- ΑΛΛΑΓΗ
            'shipping_postal_code': customer_obj.postal_code, # <-- ΑΛΛΑΓΗ
            'shipping_vat_number': customer_obj.vat_number
        })
        
    return JsonResponse({'results': customers_data})

@login_required
@require_GET
def search_products_ajax(request):
    query = request.GET.get('q', '').strip()
    initial_id = request.GET.get('initial_id', None)
    results = []
    qs = Product.objects.filter(is_active=True)
    if initial_id:
        qs = qs.filter(pk=initial_id)
    elif query:
        norm_query = normalize_text(query)
        qs = qs.filter(Q(name_normalized__icontains=norm_query) | Q(code__icontains=query) | Q(barcode__icontains=query)).distinct()
    
    products = qs.order_by('name')[:15]
    for p in products:
        results.append({
            'id': p.pk,
            'text': f"{p.name} ({p.code})",
            'price': float(p.price),
            'stock': p.stock_quantity,
            'unit': p.get_unit_of_measurement_display(),
            'unit_key': p.unit_of_measurement,
            'vat': float(p.vat_percentage) # <<< ΝΕΑ ΠΡΟΣΘΗΚΗ
        })
    return JsonResponse({'results': results})

@require_GET
def get_product_price_ajax(request):
    product_id = request.GET.get('product_id')
    if not product_id:
        return JsonResponse({'error': 'Product ID is required'}, status=400)
    try:
        product_id_int = int(product_id)
        product = Product.objects.get(pk=product_id_int)
        return JsonResponse({'price': float(product.price)})
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found', 'price': 0.00}, status=404)
    except ValueError:
        return JsonResponse({'error': 'Invalid product ID format'}, status=400)

@login_required
@require_POST
def customer_quick_add_ajax(request):
    form = CustomerForm(request.POST)
    if form.is_valid():
        customer = form.save()
        return JsonResponse({
            'success': True,
            'customer_id': customer.pk,
            'customer_text': str(customer)
        })
    else:
        return JsonResponse({'success': False, 'errors': form.errors.as_json()}, status=400)

# --- VIEW ΓΙΑ ΛΙΣΤΑ ΠΑΡΑΛΑΒΩΝ ΑΠΟΘΕΜΑΤΟΣ (ΕΝΗΜΕΡΩΜΕΝΗ) ---
@login_required
def stock_receipt_list(request):
    query = request.GET.get('q', '').strip()
    product_filter_id = request.GET.get('product_filter', None)
    date_from_str = request.GET.get('date_from', None) # Νέα παράμετρος: Ημερομηνία Από (ως string)
    date_to_str = request.GET.get('date_to', None)     # Νέα παράμετρος: Ημερομηνία Έως (ως string)

    receipts = StockReceipt.objects.select_related('product', 'user_who_recorded').all()

    # Φίλτρο βάσει επιλεγμένου προϊόντος
    if product_filter_id:
        try:
            product_filter_id_int = int(product_filter_id)
            receipts = receipts.filter(product_id=product_filter_id_int)
        except ValueError:
            product_filter_id = None 

    # Φίλτρο βάσει Ημερομηνίας Από
    if date_from_str:
        try:
            date_from_obj = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
            receipts = receipts.filter(date_received__date__gte=date_from_obj)
        except ValueError:
            date_from_str = None # Αν η ημερομηνία δεν είναι έγκυρη, την αγνοούμε για το φίλτρο

    # Φίλτρο βάσει Ημερομηνίας Έως
    if date_to_str:
        try:
            date_to_obj = datetime.datetime.strptime(date_to_str, '%Y-%m-%d').date()
            receipts = receipts.filter(date_received__date__lte=date_to_obj)
        except ValueError:
            date_to_str = None # Αν η ημερομηνία δεν είναι έγκυρη, την αγνοούμε για το φίλτρο
            
    # Φίλτρο βάσει αναζήτησης κειμένου
    if query:
        normalized_query = normalize_text(query) # Βεβαιώσου ότι η normalize_text είναι διαθέσιμη
        receipts = receipts.filter(
            Q(product__name_normalized__icontains=normalized_query if hasattr(Product, 'name_normalized') else Q(product__name__icontains=query)) |
            Q(notes__icontains=query) |
            Q(user_who_recorded__username__icontains=query)
        ).distinct()

    # Παίρνουμε όλα τα ενεργά προϊόντα για το dropdown του φίλτρου
    products_for_filter = Product.objects.filter(is_active=True).order_by('name')
    
    # Σελιδοποίηση
    paginator = Paginator(receipts, 25) # Εμφάνιση 25 εγγραφών ανά σελίδα
    page_number = request.GET.get('page')
    try:
        receipts_page = paginator.page(page_number)
    except PageNotAnInteger:
        receipts_page = paginator.page(1)
    except EmptyPage:
        receipts_page = paginator.page(paginator.num_pages)

    context = {
        'receipts': receipts_page, # Στέλνουμε το αντικείμενο της σελίδας στο template
        'query': query,
        'products_for_filter': products_for_filter,
        'product_filter_id': product_filter_id,
        'date_from_value': date_from_str, # Στέλνουμε τις string τιμές για να ξαναγεμίσουν τα πεδία
        'date_to_value': date_to_str,
        'title': 'Λίστα Παραλαβών Αποθέματος'
    }
    return render(request, 'core/stock_receipt_list.html', context)
# --- VIEW ΓΙΑ ΚΑΤΑΧΩΡΗΣΗ ΝΕΑΣ ΠΑΡΑΛΑΒΗΣ ΑΠΟΘΕΜΑΤΟΣ ---
@login_required
def receive_stock_view(request):
    """
    Λειτουργεί ως κέντρο παραλαβών:
    1. Εμφανίζει μια φόρμα για γρήγορη, χειροκίνητη εισαγωγή αποθέματος.
    2. Εμφανίζει μια λίστα με τις ανοιχτές Εντολές Αγοράς για παραλαβή βάσει παραστατικού.
    3. Χειρίζεται την υποβολή (POST) της φόρμας χειροκίνητης εισαγωγής.
    """
    
    # Χειρισμός της φόρμας χειροκίνητης εισαγωγής
    if request.method == 'POST':
        # Δημιουργούμε την instance της φόρμας με τα δεδομένα που υποβλήθηκαν
        manual_form = StockReceiptForm(request.POST)
        if manual_form.is_valid():
            # Αποθηκεύουμε τη φόρμα, αλλά δεν την κάνουμε commit στη βάση ακόμα
            receipt = manual_form.save(commit=False)
            # Ορίζουμε τον χρήστη που έκανε την καταχώρηση
            receipt.user_who_recorded = request.user
            # Τώρα αποθηκεύουμε οριστικά. Το save() του μοντέλου θα ενημερώσει το απόθεμα.
            receipt.save()
            
            product_name = receipt.product.name
            quantity_added = receipt.quantity_added
            messages.success(request, f"Η παραλαβή ποσότητας {quantity_added} για το προϊόν '{product_name}' καταχωρήθηκε επιτυχώς.")
            # Κάνουμε redirect στην ίδια σελίδα για να καθαρίσει η φόρμα
            return redirect('receive_stock')
    else:
        # Αν δεν είναι POST, δημιουργούμε μια κενή φόρμα
        manual_form = StockReceiptForm()

    # Ανεξάρτητα από το GET ή το POST, πάντα χρειαζόμαστε τη λίστα με τις ανοιχτές εντολές
    open_pos = PurchaseOrder.objects.filter(
        status__in=[PurchaseOrder.Status.ORDERED, PurchaseOrder.Status.PARTIALLY_RECEIVED]
    ).select_related('supplier').order_by('-order_date')
    
    title = 'Εισαγωγή & Παραλαβή Αποθεμάτων'
    context = {
        'title': title,
        'open_pos': open_pos,
        'manual_form': manual_form # Περνάμε τη φόρμα (είτε κενή είτε με λάθη) στο template
    }
    return render(request, 'core/receive_stock.html', context)


@login_required
def receive_po_view(request, po_pk):
    """
    Διαχειρίζεται τη φόρμα παραλαβής για μια συγκεκριμένη Εντολή Αγοράς.
    """
    po = get_object_or_404(PurchaseOrder.objects.select_related('supplier'), pk=po_pk)

    # Έλεγχος ασφαλείας: Δεν επιτρέπουμε παραλαβή για ολοκληρωμένες ή ακυρωμένες εντολές
    if po.status in [PurchaseOrder.Status.COMPLETED, PurchaseOrder.Status.CANCELLED]:
        messages.warning(request, f"Η Εντολή Αγοράς {po.po_number} είναι σε κατάσταση '{po.get_status_display()}' και δεν μπορεί να παραληφθεί.")
        return redirect('purchase_order_detail', pk=po.pk)

    # Παίρνουμε τα είδη της Εντολής Αγοράς
    po_items = po.items.select_related('product').all()

    if request.method == 'POST':
        formset = ReceivePOItemFormSet(request.POST)

        if formset.is_valid():
            try:
                with transaction.atomic():
                    items_received_in_this_transaction = False
                    
                    for form_data in formset.cleaned_data:
                        quantity_to_receive = form_data.get('quantity_to_receive', Decimal('0.00'))

                        if quantity_to_receive > 0:
                            items_received_in_this_transaction = True
                            
                            # Παίρνουμε το αρχικό είδος της παραγγελίας από τη βάση
                            po_item = PurchaseOrderItem.objects.get(pk=form_data['purchase_order_item_id'])
                            
                            # Έλεγχος για να μην ξεπεράσουμε την παραγγελθείσα ποσότητα
                            if quantity_to_receive + po_item.quantity_received > po_item.quantity:
                                messages.error(request, f"Σφάλμα στο προϊόν '{po_item.product.name}': Η ποσότητα παραλαβής ({quantity_to_receive}) υπερβαίνει την παραγγελθείσα ποσότητα που απομένει.")
                                # Ακυρώνουμε τη συναλλαγή
                                transaction.set_rollback(True)
                                # Ξαναφορτώνουμε τα αρχικά δεδομένα για να μην χαθούν οι σωστές τιμές
                                initial_data_for_formset = [
                                    {
                                        'purchase_order_item_id': item.pk,
                                        'product_name': item.product.name,
                                        'quantity_ordered': item.quantity,
                                        'quantity_already_received': item.quantity_received,
                                        'quantity_to_receive': item.quantity - item.quantity_received # Προ-συμπλήρωση με την ποσότητα που απομένει
                                    }
                                    for item in po_items
                                ]
                                formset = ReceivePOItemFormSet(initial=initial_data_for_formset)
                                return render(request, 'core/receive_po_form.html', {'po': po, 'formset': formset, 'title': f"Παραλαβή Εντολής Αγοράς {po.po_number}"})

                            # 1. Δημιουργία εγγραφής παραλαβής (το save() της θα ενημερώσει το απόθεμα)
                            StockReceipt.objects.create(
                                product=po_item.product,
                                quantity_added=quantity_to_receive,
                                user_who_recorded=request.user,
                                purchase_order_item=po_item,
                                notes=f"Παραλαβή από Εντολή Αγοράς {po.po_number}"
                            )
                            
                            # 2. Ενημέρωση της ποσότητας που παραλήφθηκε στο είδος της Εντολής Αγοράς
                            po_item.quantity_received += quantity_to_receive
                            po_item.save(update_fields=['quantity_received'])

                    if not items_received_in_this_transaction:
                        messages.warning(request, "Δεν καταχωρήσατε ποσότητα παραλαβής για κανένα προϊόν.")
                        return redirect('receive_po', po_pk=po.pk)

                    # 3. Ενημέρωση της κατάστασης της Εντολής Αγοράς
                    # Επανελέγχουμε τα items για να έχουμε τις τελευταίες τιμές
                    all_items_final = po.items.all()
                    if all(item.quantity_received >= item.quantity for item in all_items_final):
                        po.status = PurchaseOrder.Status.COMPLETED
                    else:
                        po.status = PurchaseOrder.Status.PARTIALLY_RECEIVED
                    po.save(update_fields=['status'])

                    messages.success(request, f"Η παραλαβή για την Εντολή Αγοράς {po.po_number} καταχωρήθηκε επιτυχώς.")
                    return redirect('purchase_order_detail', pk=po.pk)

            except Exception as e:
                messages.error(request, f"Προέκυψε ένα μη αναμενόμενο σφάλμα: {e}")

        # Αν το formset δεν είναι valid, ξαναεμφανίζουμε τη σελίδα με τα σφάλματα
        return render(request, 'core/receive_po_form.html', {'po': po, 'formset': formset, 'title': f"Παραλαβή Εντολής Αγοράς {po.po_number}"})

    else: # GET request
        # Προετοιμασία των αρχικών δεδομένων για το formset
        initial_data_for_formset = [
            {
                'purchase_order_item_id': item.pk,
                'product_name': f"{item.product.name} ({item.product.code})",
                'quantity_ordered': item.quantity,
                'quantity_already_received': item.quantity_received,
                'quantity_to_receive': item.quantity - item.quantity_received # Προ-συμπλήρωση με την ποσότητα που απομένει
            }
            for item in po_items
        ]
        
        formset = ReceivePOItemFormSet(initial=initial_data_for_formset)
        
        context = {
            'po': po,
            'formset': formset,
            'title': f"Παραλαβή Εντολής Αγοράς {po.po_number}"
        }
        return render(request, 'core/receive_po_form.html', context)
@login_required
def activity_log_list_view(request):
    log_entries_list = ActivityLog.objects.select_related('user', 'content_type').all().order_by('-action_time')

    user_filter_id = request.GET.get('user_filter', None) # Τώρα θα είναι ID
    action_type_filter = request.GET.get('action_type_filter', None)
    
    # Φίλτρο βάσει επιλεγμένου χρήστη (με ID)
    if user_filter_id:
        try:
            selected_user_id = int(user_filter_id)
            log_entries_list = log_entries_list.filter(user_id=selected_user_id)
        except ValueError:
            user_filter_id = None # Αγνοούμε μη έγκυρες τιμές

    if action_type_filter:
        log_entries_list = log_entries_list.filter(action_type=action_type_filter)

    paginator = Paginator(log_entries_list, 25)
    page_number = request.GET.get('page')
    try:
        log_entries = paginator.page(page_number)
    except PageNotAnInteger:
        log_entries = paginator.page(1)
    except EmptyPage:
        log_entries = paginator.page(paginator.num_pages)

    all_users = User.objects.all().order_by('username') # Παίρνουμε όλους τους χρήστες για το dropdown

    context = {
        'log_entries': log_entries,
        'title': 'Ιστορικό Ενεργειών Συστήματος',
        'ACTION_TYPE_CHOICES': ActivityLog.ACTION_TYPE_CHOICES,
        'all_users': all_users, # <<< ΝΕΟ: Λίστα χρηστών για το φίλτρο
        'user_filter_id_value': int(user_filter_id) if user_filter_id and user_filter_id.isdigit() else None, # <<< ΝΕΟ: Το επιλεγμένο ID για το template
        'action_type_filter_value': action_type_filter if action_type_filter else '',
    }
    return render(request, 'core/activity_log_list.html', context)
def custom_logout_view(request):
    if request.user.is_authenticated: # Προαιρετικός έλεγχος
        auth_logout_function(request)
        messages.success(request, "Έχετε αποσυνδεθεί επιτυχώς.") # Προαιρετικό μήνυμα
    return redirect('home')
@login_required
@require_POST # Διασφαλίζει ότι αυτή η view καλείται μόνο με POST request
def stock_receipt_delete(request, pk):
    receipt = get_object_or_404(StockReceipt, pk=pk)
    product_name_for_message = receipt.product.name if receipt.product else "άγνωστου προϊόντος"
    receipt_quantity_for_message = receipt.quantity_added
    
    try:
        receipt.delete() # Η διαγραφή θα ενεργοποιήσει τον post_delete signal handler
                         # που θα χειριστεί την αφαίρεση αποθέματος και το ActivityLog.
        messages.success(request, 
                         f"Η παραλαβή ποσότητας {receipt_quantity_for_message} για το προϊόν '{product_name_for_message}' διαγράφηκε επιτυχώς.")
    except Exception as e:
        messages.error(request, f"Προέκυψε σφάλμα κατά τη διαγραφή της παραλαβής: {e}")
        # Εδώ θα μπορούσες να προσθέσεις logging του σφάλματος e αν χρειαστεί
        
    return redirect('stock_receipt_list')
@staff_member_required # Μόνο staff μέλη μπορούν να δουν αυτή τη σελίδα
@login_required # Για επιπλέον σιγουριά, αν και το staff_member_required το καλύπτει
def user_list_view(request):
    # Παίρνουμε όλους τους χρήστες, με τους νεότερους πρώτα
    users_qs = User.objects.all().order_by('-date_joined')
    
    # --- ΝΕΑ ΛΟΓΙΚΗ ΓΙΑ ΣΕΛΙΔΟΠΟΙΗΣΗ ---
    paginator = Paginator(users_qs, 20) # Π.χ., 20 χρήστες ανά σελίδα
    page_number = request.GET.get('page')
    try:
        users_page = paginator.page(page_number)
    except PageNotAnInteger:
        users_page = paginator.page(1)
    except EmptyPage:
        users_page = paginator.page(paginator.num_pages)
    
    # --- ΤΕΛΟΣ ΝΕΑΣ ΛΟΓΙΚΗΣ ---

    context = {
        'users': users_page, # <<< ΑΛΛΑΓΗ: Περνάμε το page object
        'title': 'Διαχείριση Χρηστών'
    }
    return render(request, 'core/user_list.html', context)
@staff_member_required
@login_required
def user_create_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            new_user = form.save()
            # Προαιρετικά: Αν θέλεις να προσθέσεις τον νέο χρήστη σε κάποια default ομάδα
            # group = Group.objects.get(name='YourDefaultGroupName') 
            # new_user.groups.add(group)
            messages.success(request, f"Ο χρήστης '{new_user.username}' δημιουργήθηκε επιτυχώς!")
            return redirect('user_list_view') # Ανακατεύθυνση στη λίστα χρηστών
        else:
            messages.error(request, "Προέκυψαν σφάλματα. Παρακαλώ διορθώστε τα παρακάτω.")
    else:
        form = CustomUserCreationForm()

    context = {
        'form': form,
        'title': 'Δημιουργία Νέου Χρήστη'
    }
    return render(request, 'core/user_create.html', context)
@staff_member_required
@login_required
def user_edit_view(request, pk):
    user_to_edit = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=user_to_edit)
        if form.is_valid():
            updated_user = form.save() # Η αποθήκευση θα ενεργοποιήσει τον log_user_save signal handler
            messages.success(request, f"Τα στοιχεία του χρήστη '{updated_user.username}' ενημερώθηκαν επιτυχώς!")
            return redirect('user_list_view') # Ανακατεύθυνση στη λίστα χρηστών
        else:
            messages.error(request, "Προέκυψαν σφάλματα. Παρακαλώ διορθώστε τα παρακάτω.")
    else:
        form = CustomUserChangeForm(instance=user_to_edit)

    context = {
        'form': form,
        'user_to_edit': user_to_edit, # Για να μπορούμε να εμφανίσουμε π.χ. το username στον τίτλο
        'title': f'Επεξεργασία Χρήστη: {user_to_edit.username}'
    }
    return render(request, 'core/user_edit.html', context)

@login_required
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    
    # Λογική για τη φόρμα επισύναψης αρχείων
    attachment_form = AttachmentForm()
    if request.method == 'POST' and 'upload_file' in request.POST: # Ελέγχουμε αν υποβλήθηκε η φόρμα αρχείων
        attachment_form = AttachmentForm(request.POST, request.FILES)
        if attachment_form.is_valid():
            attachment = attachment_form.save(commit=False)
            attachment.uploaded_by = request.user
            attachment.content_object = customer # Σύνδεση με τον πελάτη
            attachment.save()
            messages.success(request, f"Το αρχείο '{attachment.file.name}' ανέβηκε επιτυχώς.")
            return redirect('customer_detail', pk=customer.pk)

    # Ανάκτηση όλων των σχετικών εγγραφών
    branches = customer.branches.all()
    query = request.GET.get('q', '').strip()

    customer_orders = customer.orders.all().order_by('-order_date', '-pk')
    customer_invoices = customer.invoices.all().order_by('-issue_date', '-pk')
    customer_delivery_notes = customer.delivery_notes.all().order_by('-issue_date', '-pk')
    
    # Ανάκτηση των επισυναπτόμενων για τον συγκεκριμένο πελάτη
    customer_content_type = ContentType.objects.get_for_model(Customer)
    attachments = Attachment.objects.filter(content_type=customer_content_type, object_id=customer.pk)

    if query:
        customer_orders = customer_orders.filter(order_number__icontains=query)
        customer_invoices = customer_invoices.filter(invoice_number__icontains=query)
        customer_delivery_notes = customer_delivery_notes.filter(delivery_note_number__icontains=query)

    completed_orders_value = customer.orders.filter(status=Order.STATUS_COMPLETED).aggregate(
        total_value=Sum('total_amount')
    )['total_value'] or 0.00

    context = {
        'customer': customer,
        'branches': branches,
        'customer_orders': customer_orders,
        'customer_invoices': customer_invoices,
        'customer_delivery_notes': customer_delivery_notes,
        'completed_orders_value': completed_orders_value,
        'query': query,
        'title': f'Καρτέλα Πελάτη: {customer.get_full_name()}',
        'attachments': attachments, # Προσθήκη των αρχείων στο context
        'attachment_form': attachment_form, # Προσθήκη της φόρμας στο context
    }
    return render(request, 'core/customer_detail.html', context)

@login_required
def payment_create_view(request, customer_pk=None):
    customer = None
    if customer_pk:
        customer = get_object_or_404(Customer, pk=customer_pk)
    else:
        customer_id_from_get = request.GET.get('customer_id')
        if customer_id_from_get:
            try:
                customer = Customer.objects.get(pk=customer_id_from_get)
            except (Customer.DoesNotExist, ValueError):
                messages.error(request, "Επιλέχθηκε μη έγκυρος πελάτης.")
                return redirect('payment_create')

    if request.method == 'POST':
        form = PaymentForm(request.POST, customer_instance=customer)
        if form.is_valid():
            payment = form.save(commit=False)
            # Το customer πρέπει να οριστεί εδώ, καθώς η φόρμα δεν το ξέρει αν το πεδίο είναι disabled/hidden.
            payment.customer = customer
            payment.recorded_by = request.user
            
            payment.save()
            form.save_m2m() # Αποθηκεύει τη σχέση με τα τιμολόγια
            
            messages.success(request, f"Η πληρωμή ποσού {payment.amount_paid}€ καταχωρήθηκε.")
            return redirect('customer_financial_detail', pk=payment.customer.pk)
    else:
        # --- Η ΔΙΟΡΘΩΣΗ ΕΙΝΑΙ ΕΔΩ ---
        # Πρέπει να περάσουμε την αρχική τιμή για το κρυφό πεδίο 'customer'
        initial_data = {}
        if customer:
            initial_data['customer'] = customer.pk
        form = PaymentForm(initial=initial_data, customer_instance=customer)
        # --- ΤΕΛΟΣ ΔΙΟΡΘΩΣΗΣ ---

    outstanding_invoices = form.fields['invoices'].queryset if 'invoices' in form.fields else Invoice.objects.none()
    
    context = {
        'form': form,
        'customer': customer,
        'outstanding_invoices': outstanding_invoices,
        'title': f"Καταχώρηση Νέας Πληρωμής{' για ' + customer.get_full_name() if customer else ''}"
    }
    return render(request, 'core/payment_form.html', context)

@login_required
def customer_financial_detail_view(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    today = timezone.now().date()

    all_invoices = customer.invoices.filter(status__in=[Invoice.STATUS_ISSUED, Invoice.STATUS_PAID])
    all_active_payments = customer.payments.filter(status=Payment.STATUS_ACTIVE)
    all_issued_credit_notes = customer.credit_notes.filter(status=CreditNote.Status.ISSUED)

    total_invoiced = all_invoices.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    total_paid = all_active_payments.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
    
    overdue_invoices = all_invoices.filter(status=Invoice.STATUS_ISSUED, due_date__lt=today)
    overdue_count = overdue_invoices.count()
    overdue_amount = sum(inv.outstanding_amount for inv in overdue_invoices)
    
    credit_percentage = 0
    if customer.credit_limit and customer.credit_limit > 0:
        balance_for_calc = customer.balance if customer.balance > 0 else 0
        credit_percentage = (balance_for_calc / customer.credit_limit) * 100

    transaction_log = []

    # Προσθέτουμε τα τιμολόγια (Χρεώσεις)
    for invoice in all_invoices:
        
        # --- ΝΕΑ ΔΙΟΡΘΩΜΕΝΗ ΛΟΓΙΚΗ ---
        # Φτιάχνουμε την περιγραφή ξεχωριστά
        description = 'Τιμολόγιο Πώλησης'
        if invoice.delivery_note and invoice.delivery_note.customer.is_branch:
            branch_customer = invoice.delivery_note.customer
            description = f"Τιμολόγηση Δ.Α. προς: {branch_customer.company_name or branch_customer.get_full_name()}"

        transaction_log.append({
            'date': invoice.issue_date,
            'doc_type': 'Τιμολόγιο', # <-- Το κρατάμε σταθερό για τη λογική του link
            'description': description, # <-- Περνάμε την αναλυτική περιγραφή ξεχωριστά
            'doc_number': invoice.invoice_number,
            'debit': invoice.total_amount,
            'credit': Decimal('0.00'),
            'obj': invoice,
            'is_overdue': invoice.is_overdue
        })
        
    # Προσθέτουμε τις πληρωμές (Πιστώσεις)
    for payment in all_active_payments:
         transaction_log.append({
            'date': payment.payment_date,
            'doc_type': 'Πληρωμή',
            'description': f"Πληρωμή ({payment.get_payment_method_display()})",
            'doc_number': payment.receipt_number,
            'debit': Decimal('0.00'),
            'credit': payment.amount_paid,
            'obj': payment,
            'is_overdue': False
        })

    # Προσθέτουμε τα Πιστωτικά (Πιστώσεις)
    for cn in all_issued_credit_notes:
        transaction_log.append({
            'date': cn.issue_date,
            'doc_type': 'Πιστωτικό Τιμολόγιο',
            'description': 'Πιστωτικό Τιμολόγιο',
            'doc_number': cn.credit_note_number,
            'debit': Decimal('0.00'),
            'credit': cn.total_amount,
            'obj': cn,
            'is_overdue': False
        })
        
    transaction_log.sort(key=lambda x: x['date'])

    running_balance = Decimal('0.00')
    for entry in transaction_log:
        running_balance += entry['debit'] - entry['credit']
        entry['running_balance'] = running_balance

    context = {
        'customer': customer,
        'title': f"Οικονομική Καρτέλα: {customer.get_full_name()}",
        'total_invoiced': total_invoiced,
        'total_paid': total_paid,
        'overdue_count': overdue_count,
        'overdue_amount': overdue_amount,
        'credit_percentage': credit_percentage,
        'transaction_log': transaction_log,
    }
    
    return render(request, 'core/customer_financial_detail.html', context)
@login_required
def payment_cancel_view(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    
    # Έλεγχος αν η πληρωμή είναι ήδη ακυρωμένη ή αν δεν ανήκει σε πελάτη (ασφάλεια)
    if payment.status == Payment.STATUS_CANCELLED:
        messages.warning(request, f"Η πληρωμή {payment.receipt_number} είναι ήδη ακυρωμένη.")
        return redirect('customer_financial_detail', pk=payment.customer.pk)
    
    if not payment.customer: # Ασφάλεια, κάθε πληρωμή πρέπει να έχει πελάτη
        messages.error(request, "Σφάλμα: Η πληρωμή δεν έχει συσχετισμένο πελάτη.")
        return redirect('home') # ή όπου αλλού είναι κατάλληλο

    if request.method == 'POST':
        form = PaymentCancellationForm(request.POST)
        if form.is_valid():
            payment.status = Payment.STATUS_CANCELLED
            payment.cancellation_reason = form.cleaned_data['cancellation_reason']
            payment.cancelled_at = timezone.now()
            payment.cancelled_by = request.user
            payment.save() # Αυτό θα ενεργοποιήσει το signal για το ActivityLog και την ενημέρωση του balance

            messages.success(request, f"Η πληρωμή {payment.receipt_number} ακυρώθηκε επιτυχώς.")
            return redirect('customer_financial_detail', pk=payment.customer.pk)
    else:
        form = PaymentCancellationForm()

    context = {
        'form': form,
        'payment': payment,
        'customer': payment.customer,
        'title': f"Ακύρωση Πληρωμής {payment.receipt_number}"
    }
    return render(request, 'core/payment_cancel_form.html', context)
@login_required
def payment_edit_view(request, pk):
    payment_to_edit = get_object_or_404(Payment, pk=pk)

    if payment_to_edit.status == Payment.STATUS_CANCELLED:
        messages.error(request, f"Η πληρωμή {payment_to_edit.receipt_number} είναι ακυρωμένη και δεν μπορεί να επεξεργαστεί.")
        if payment_to_edit.customer:
            return redirect('customer_financial_detail', pk=payment_to_edit.customer.pk)
        return redirect('home') # Fallback αν δεν υπάρχει customer

    if request.method == 'POST':
        # Περνάμε is_edit_mode=True για να απενεργοποιηθούν τα σωστά πεδία
        form = PaymentForm(request.POST, instance=payment_to_edit, is_edit_mode=True)
        if form.is_valid():
            # Πριν την αποθήκευση, καταγράφουμε τις αλλαγές για το ActivityLog
            changed_fields_desc = []
            if form.changed_data: # Λίστα με τα ονόματα των πεδίων που άλλαξαν
                for field_name in form.changed_data:
                    # Παίρνουμε το verbose_name του πεδίου για πιο φιλικό μήνυμα
                    field_label = payment_to_edit._meta.get_field(field_name).verbose_name
                    changed_fields_desc.append(f"'{field_label}'")
            
            payment = form.save(commit=False)
            # Το recorded_by δεν αλλάζει κατά την επεξεργασία, παραμένει ο αρχικός.
            # Αν θέλουμε να καταγράφουμε ποιος έκανε την τελευταία επεξεργασία,
            # θα χρειαζόταν ένα πεδίο 'last_edited_by' στο μοντέλο Payment.

            # Προσθέτουμε τις λεπτομέρειες αλλαγών στο instance για να τις πιάσει το signal
            if changed_fields_desc:
                payment._change_details_for_log = f"Έγιναν αλλαγές στα πεδία: {', '.join(changed_fields_desc)}."
            else:
                payment._change_details_for_log = "Δεν έγιναν αλλαγές σε πεδία της φόρμας (πιθανόν αποθήκευση χωρίς τροποποιήσεις)."


            payment.save() # Το save του Payment θα χειριστεί το balance αν άλλαζε το amount_paid (που δεν αλλάζει εδώ)
                           # και θα ενεργοποιήσει το post_save signal.
            
            messages.success(request, f"Οι αλλαγές στην πληρωμή {payment_to_edit.receipt_number} αποθηκεύτηκαν.")
            if payment_to_edit.customer:
                return redirect('customer_financial_detail', pk=payment_to_edit.customer.pk)
            return redirect('home') # Fallback
    else:
        # Περνάμε is_edit_mode=True
        form = PaymentForm(instance=payment_to_edit, customer_instance=payment_to_edit.customer, is_edit_mode=True)

    context = {
        'form': form,
        'payment': payment_to_edit,
        'customer': payment_to_edit.customer,
        'is_edit_mode': True, # Για να ξέρει το template ότι είναι επεξεργασία
        'title': f"Επεξεργασία Πληρωμής {payment_to_edit.receipt_number}"
    }
    return render(request, 'core/payment_form.html', context) # Μπορούμε να επαναχρησιμοποιήσουμε το payment_form.html
@login_required
def view_payment_receipt_pdf(request, pk): # Άλλαξα το όνομα για να είναι σαφές ότι είναι για PDF
    if not WEASYPRINT_AVAILABLE:
        # Εναλλακτικά, θα μπορούσες να εμφανίσεις μια απλή HTML σελίδα εδώ
        # ή ένα μήνυμα ότι η δημιουργία PDF δεν είναι διαθέσιμη.
        # For now, let's raise a 404 or show an error message.
        # messages.error(request, "Η βιβλιοθήκη WeasyPrint δεν είναι εγκατεστημένη. Η δημιουργία PDF δεν είναι δυνατή.")
        # return redirect(request.META.get('HTTP_REFERER', 'home')) # Επιστροφή στην προηγούμενη σελίδα
        raise Http404("Η λειτουργία δημιουργίας PDF δεν είναι διαθέσιμη. Η βιβλιοθήκη WeasyPrint λείπει.")

    payment = get_object_or_404(Payment, pk=pk)
    
    # Αν η πληρωμή είναι ακυρωμένη, ίσως δεν θέλουμε να εκδίδουμε απόδειξη
    # ή η απόδειξη θα πρέπει να το αναφέρει καθαρά.
    # Προς το παρόν, ας την επιτρέπουμε, αλλά μπορείς να προσθέσεις έλεγχο εδώ.
    if payment.status == Payment.STATUS_CANCELLED:
        messages.error(request, "Δεν είναι δυνατή η έκδοση απόδειξης για ακυρωμένη πληρωμή.")
        if payment.customer:
            return redirect('customer_financial_detail', pk=payment.customer.pk)
        return redirect('home')

    company_info_from_settings = getattr(settings, 'COMPANY_INFO', {})

    context = {
        'payment': payment,
        'company_info': company_info_from_settings,
    }
    
    # Render το HTML template σε string
    html_string = render_to_string('core/payment_receipt_pdf.html', context)
    
    # Δημιουργία του PDF με WeasyPrint
    # Το base_url είναι σημαντικό αν το HTML template σου έχει εξωτερικούς πόρους (π.χ. εικόνες, CSS αρχεία)
    # που είναι προσβάσιμοι μέσω URL. Για απλό HTML/CSS μέσα στο template, μπορείς να το παραλείψεις.
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/')) # Αν χρειάζεται base_url
    html = HTML(string=html_string)
    
    # Παράγουμε το PDF ως bytes
    pdf_bytes = html.write_pdf()
    
    # Δημιουργούμε το HttpResponse
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="receipt_{payment.receipt_number or payment.pk}.pdf"'
    # Αντί για 'inline', μπορείς να χρησιμοποιήσεις 'attachment' για να γίνει απευθείας download:
    # response['Content-Disposition'] = f'attachment; filename="receipt_{payment.receipt_number or payment.pk}.pdf"'
    
    return response
@login_required
@require_GET # Αυτή η view θα δέχεται μόνο GET requests
def get_customer_orders_ajax(request, customer_id):
    try:
        customer = Customer.objects.get(pk=customer_id)
        # Φέρνουμε παραγγελίες που μπορεί να είναι σχετικές για πληρωμή
        # (π.χ. όχι ακυρωμένες). Μπορείς να προσαρμόσεις τα status αν χρειάζεται.
        orders = Order.objects.filter(
            customer=customer,
            status__in=[Order.STATUS_PENDING, Order.STATUS_PROCESSING, Order.STATUS_COMPLETED]
        ).order_by('-order_date', '-pk')
        
        orders_data = [{
            'id': order.pk,
            'text': f"{order.order_number} ({order.order_date.strftime('%d/%m/%Y')}) - {order.total_amount}€ [{order.get_status_display()}]"
        } for order in orders]
        
        return JsonResponse({'results': orders_data})
    except Customer.DoesNotExist:
        return JsonResponse({'results': [], 'error': 'Customer not found'}, status=404)
    except Exception as e:
        # Καλό είναι να κάνεις log το σφάλμα e εδώ για debugging
        return JsonResponse({'results': [], 'error': 'Server error'}, status=500)
@login_required
def all_payments_list_view(request):
    payments_qs = Payment.objects.select_related(
        'customer', 'order', 'recorded_by', 'cancelled_by'
    ).all().order_by('-payment_date', '-receipt_number')

    # Λήψη τιμών φίλτρων από το GET request
    query_text = request.GET.get('q', '').strip()
    customer_filter_id = request.GET.get('customer_filter', None)
    status_filter = request.GET.get('status_filter', 'all') # 'all', 'active', 'cancelled'
    method_filter = request.GET.get('method_filter', 'all')
    
    date_from_str = request.GET.get('date_from', None)
    date_to_str = request.GET.get('date_to', None)
    
    value_date_from_str = request.GET.get('value_date_from', None)
    value_date_to_str = request.GET.get('value_date_to', None)

    # Εφαρμογή φίλτρων
    if query_text:
        payments_qs = payments_qs.filter(
            Q(receipt_number__icontains=query_text) |
            Q(reference_number__icontains=query_text) |
            Q(notes__icontains=query_text) # Προσθήκη αναζήτησης και στις σημειώσεις
        )

    if customer_filter_id:
        try:
            payments_qs = payments_qs.filter(customer_id=int(customer_filter_id))
        except ValueError:
            customer_filter_id = None # Αγνοούμε μη έγκυρη τιμή

    if status_filter != 'all':
        payments_qs = payments_qs.filter(status=status_filter)

    if method_filter != 'all':
        payments_qs = payments_qs.filter(payment_method=method_filter)

    # Φίλτρο Ημερομηνίας Πληρωμής
    if date_from_str:
        try:
            date_from_obj = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
            payments_qs = payments_qs.filter(payment_date__gte=date_from_obj)
        except ValueError:
            date_from_str = None 
    if date_to_str:
        try:
            date_to_obj = datetime.datetime.strptime(date_to_str, '%Y-%m-%d').date()
            payments_qs = payments_qs.filter(payment_date__lte=date_to_obj)
        except ValueError:
            date_to_str = None

    # Φίλτρο Ημερομηνίας Λήξης/Value Date
    if value_date_from_str:
        try:
            value_date_from_obj = datetime.datetime.strptime(value_date_from_str, '%Y-%m-%d').date()
            payments_qs = payments_qs.filter(value_date__gte=value_date_from_obj)
        except ValueError:
            value_date_from_str = None
    if value_date_to_str:
        try:
            value_date_to_obj = datetime.datetime.strptime(value_date_to_str, '%Y-%m-%d').date()
            payments_qs = payments_qs.filter(value_date__lte=value_date_to_obj)
        except ValueError:
            value_date_to_str = None
            
    # Υπολογισμός συνόλου ενεργών πληρωμών (μετά τα φίλτρα)
    # Αυτό το σύνολο αφορά μόνο τις πληρωμές που εμφανίζονται στην τρέχουσα φιλτραρισμένη λίστα, όχι όλες τις πληρωμές.
    # Αν θέλεις το σύνολο όλων των ενεργών πληρωμών ανεξαρτήτως φίλτρων σελιδοποίησης, υπολόγισέ το πριν τη σελιδοποίηση.
    active_payments_in_queryset = payments_qs.filter(status=Payment.STATUS_ACTIVE)
    total_active_payments_amount = active_payments_in_queryset.aggregate(total=Sum('amount_paid'))['total'] or 0.00

    # Σελιδοποίηση
    paginator = Paginator(payments_qs, 25) # Π.χ. 25 πληρωμές ανά σελίδα
    page_number = request.GET.get('page')
    try:
        payments_page = paginator.page(page_number)
    except PageNotAnInteger:
        payments_page = paginator.page(1)
    except EmptyPage:
        payments_page = paginator.page(paginator.num_pages)

    # Δεδομένα για τα dropdown των φίλτρων
    customers_for_filter = Customer.objects.all().order_by('company_name', 'last_name', 'first_name')
    payment_method_choices = Payment.PAYMENT_METHOD_CHOICES
    payment_status_choices = Payment.PAYMENT_STATUS_CHOICES


    context = {
        'payments': payments_page,
        'total_active_payments_amount': total_active_payments_amount,
        # Τιμές φίλτρων για να ξαναγεμίσουν τα πεδία στη φόρμα
        'query_text': query_text,
        'customer_filter_id': int(customer_filter_id) if customer_filter_id and customer_filter_id.isdigit() else None,
        'status_filter': status_filter,
        'method_filter': method_filter,
        'date_from_value': date_from_str,
        'date_to_value': date_to_str,
        'value_date_from_value': value_date_from_str,
        'value_date_to_value': value_date_to_str,
        # Επιλογές για τα select των φίλτρων
        'customers_for_filter': customers_for_filter,
        'payment_method_choices': payment_method_choices,
        'payment_status_choices': payment_status_choices, # Περνάμε και τις επιλογές κατάστασης
        'title': 'Λίστα Όλων των Πληρωμών'
    }
    return render(request, 'core/all_payments_list.html', context)
@login_required
def view_order_pdf(request, pk):
    if not WEASYPRINT_AVAILABLE:
        raise Http404("Η λειτουργία δημιουργίας PDF δεν είναι διαθέσιμη. Η βιβλιοθήκη WeasyPrint λείπει.")

    order = get_object_or_404(Order.objects.select_related('customer'), pk=pk)
    order_items = order.items.select_related('product').all()

    # Λογική για τη σύνοψη ποσοτήτων ανά μονάδα μέτρησης
    quantity_summary = defaultdict(float)
    for item in order_items:
        if item.product:
            unit_display = item.product.get_unit_of_measurement_display()
            quantity_summary[unit_display] += float(item.quantity)
    
    company_info_from_settings = getattr(settings, 'COMPANY_INFO', {})

    context = {
        'order': order,
        'order_items': order_items,
        'quantity_summary': dict(quantity_summary), # Μετατροπή σε κανονικό dict για το template
        'company_info': company_info_from_settings,
    }
    
    html_string = render_to_string('core/order_pdf.html', context)
    
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    pdf_bytes = html.write_pdf()
    
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="order_{order.order_number or order.pk}.pdf"'
    
    return response
@login_required
def export_customers_to_excel(request):
    # --- Η λογική για την ανάκτηση και το φιλτράρισμα των πελατών παραμένει ακριβώς ίδια ---
    query_from_request = request.GET.get('q', None)
    customers_qs = Customer.objects.all().order_by('company_name', 'last_name')

    if query_from_request:
        query = query_from_request.strip()
        if query:
            normalized_user_query = normalize_text(query)
            customers_qs = customers_qs.filter(
                Q(first_name_normalized__icontains=normalized_user_query) |
                Q(last_name_normalized__icontains=normalized_user_query) |
                Q(company_name_normalized__icontains=normalized_user_query) |
                Q(code__icontains=query) |
                Q(email__icontains=query) |
                Q(phone__icontains=query)
            ).distinct()

    # Προετοιμασία δεδομένων για το DataFrame
    data_for_export = []
    for customer in customers_qs:
        data_for_export.append({
            'Κωδικός': customer.code,
            'Όνομα': customer.first_name,
            'Επώνυμο': customer.last_name,
            'Επωνυμία Εταιρείας': customer.company_name,
            'Email': customer.email,
            'Τηλέφωνο': customer.phone,
            'Διεύθυνση': customer.address,
            'Πόλη': customer.city,
            'Τ.Κ.': customer.postal_code,
            'Α.Φ.Μ.': customer.vat_number,
            'Δ.Ο.Υ.': customer.doy,
            'Υπόλοιπο (€)': customer.balance,
        })
    
    # Αν δεν υπάρχουν δεδομένα, επέστρεψε ένα μήνυμα
    if not data_for_export:
        messages.warning(request, "Δεν βρέθηκαν πελάτες για εξαγωγή.")
        return redirect(request.META.get('HTTP_REFERER', 'customer_list'))

    # Δημιουργία Pandas DataFrame
    df = pd.DataFrame(data_for_export)

    # Δημιουργία ενός in-memory buffer για το αρχείο Excel
    excel_buffer = io.BytesIO()

    # --- Χρήση του ExcelWriter για να μπορέσουμε να αλλάξουμε το πλάτος των στηλών ---
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Πελάτες')

        # Πρόσβαση στο worksheet του openpyxl
        worksheet = writer.sheets['Πελάτες']
        
        # Εισαγωγή του get_column_letter εδώ για να μην χρειάζεται global import
        from openpyxl.utils import get_column_letter

        # Loop μέσα από τις στήλες του DataFrame για να ορίσουμε το πλάτος
        for idx, col in enumerate(df):
            series = df[col]
            max_len = max((
                series.astype(str).map(len).max(),
                len(str(series.name))
                )) + 2  # Προσθέτουμε ένα μικρό padding
            
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len

    # --- Τέλος νέας λογικής ---

    # Επιστροφή του buffer ως HTTP response για download
    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="customers_export.xlsx"'

    return response
@login_required
def export_products_to_excel(request):
    # --- Η λογική για την ανάκτηση και το φιλτράρισμα των προϊόντων παραμένει ακριβώς ίδια ---
    query_from_request = request.GET.get('q', None)
    products_qs = Product.objects.all().order_by('name')
    if query_from_request:
        query = query_from_request.strip()
        if query:
            normalized_user_query = normalize_text(query)
            products_qs = products_qs.filter(
                Q(name_normalized__icontains=normalized_user_query) |
                Q(code__icontains=query) |
                Q(barcode__icontains=query) |
                Q(description__icontains=query)
            ).distinct()

    data_for_export = []
    for product in products_qs:
        data_for_export.append({
            'Κωδικός': product.code,
            'Barcode': product.barcode,
            'Όνομα Προϊόντος': product.name,
            'Περιγραφή': product.description,
            'Κατάσταση': 'Ενεργό' if product.is_active else 'Ανενεργό',
            'Ποσότητα Αποθέματος': product.stock_quantity,
            'Ελάχιστο Όριο Αποθέματος': product.min_stock_level,
            'Μονάδα Μέτρησης': product.get_unit_of_measurement_display(),
            'Τιμή Πώλησης (€)': product.price,
            'Τιμή Κόστους (€)': product.cost_price,
        })

    # --- Από εδώ και κάτω είναι η νέα, βελτιωμένη λογική ---
    
    # Αν δεν υπάρχουν δεδομένα, επέστρεψε ένα κενό αρχείο ή ένα μήνυμα
    if not data_for_export:
        messages.warning(request, "Δεν βρέθηκαν προϊόντα για εξαγωγή.")
        # Επιστρέφουμε στην προηγούμενη σελίδα με βάση το HTTP_REFERER
        # ή σε μια default σελίδα αν το referer δεν είναι διαθέσιμο
        return redirect(request.META.get('HTTP_REFERER', 'product_list'))

    df = pd.DataFrame(data_for_export)
    
    excel_buffer = io.BytesIO()

    # --- Χρήση του ExcelWriter για να μπορέσουμε να αλλάξουμε το πλάτος των στηλών ---
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Προϊόντα')

        # Πρόσβαση στο worksheet του openpyxl
        worksheet = writer.sheets['Προϊόντα']
        
        # Loop μέσα από τις στήλες του DataFrame για να ορίσουμε το πλάτος
        for idx, col in enumerate(df):  # df.columns για να πάρουμε τα ονόματα των στηλών
            series = df[col]
            max_len = max((
                series.astype(str).map(len).max(),  # Μέγιστο μήκος δεδομένων στη στήλη
                len(str(series.name))  # Μήκος του τίτλου της στήλης
                )) + 2  # Προσθέτουμε ένα μικρό padding
            
            # Το idx + 1 αντιστοιχεί στον αριθμό της στήλης στο Excel (που ξεκινάει από 1)
            # Μετατρέπουμε τον αριθμό σε γράμμα στήλης (π.χ. 1->A, 2->B)
            # και ορίζουμε το πλάτος.
            from openpyxl.utils import get_column_letter
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len

    # --- Τέλος νέας λογικής ---

    # Επιστροφή του buffer ως HTTP response για download
    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'

    return response
@login_required
def export_orders_to_excel(request):
    # Παίρνουμε το ίδιο queryset και τους ίδιους κανόνες φιλτραρίσματος από τη view order_list
    orders_qs = Order.objects.select_related('customer').all().order_by('-order_date', '-pk')

    # Λήψη τιμών φίλτρων από το GET request
    status_filter = request.GET.get('status', 'all')
    query_text = request.GET.get('q', '').strip()
    customer_filter_id = request.GET.get('customer_filter', None)
    date_from_str = request.GET.get('date_from', None)
    date_to_str = request.GET.get('date_to', None)

    # Εφαρμογή φίλτρων (ακριβώς όπως στη view order_list)
    if status_filter != 'all':
        orders_qs = orders_qs.filter(status=status_filter)
    if customer_filter_id:
        try:
            orders_qs = orders_qs.filter(customer_id=int(customer_filter_id))
        except (ValueError, TypeError):
            pass
    if date_from_str:
        try:
            date_from_obj = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
            orders_qs = orders_qs.filter(order_date__gte=date_from_obj)
        except ValueError:
            pass
    if date_to_str:
        try:
            date_to_obj = datetime.datetime.strptime(date_to_str, '%Y-%m-%d').date()
            orders_qs = orders_qs.filter(order_date__lte=date_to_obj)
        except ValueError:
            pass
    if query_text:
        normalized_query = normalize_text(query_text)
        orders_qs = orders_qs.filter(
            Q(order_number__icontains=query_text) |
            Q(customer__first_name_normalized__icontains=normalized_query) |
            Q(customer__last_name_normalized__icontains=normalized_query) |
            Q(customer__company_name_normalized__icontains=normalized_query) |
            Q(customer__code__icontains=query_text)
        ).distinct()
    
    # Προετοιμασία δεδομένων για το DataFrame
    data_for_export = []
    for order in orders_qs:
        data_for_export.append({
            'Αρ. Παραγγελίας': order.order_number,
            'Ημερομηνία Παραγγελίας': order.order_date,
            'Πελάτης': str(order.customer) if order.customer else '[Διαγραμμένος Πελάτης]',
            'Κατάσταση': order.get_status_display(),
            'Συνολικό Ποσό (€)': order.total_amount,
            'Ημερομηνία Παράδοσης': order.delivery_date,
            'Σχόλια': order.comments,
        })

    if not data_for_export:
        messages.warning(request, "Δεν βρέθηκαν παραγγελίες για εξαγωγή με βάση τα επιλεγμένα κριτήρια.")
        return redirect(request.META.get('HTTP_REFERER', 'order_list'))

    df = pd.DataFrame(data_for_export)

    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Παραγγελίες')
        worksheet = writer.sheets['Παραγγελίες']
        
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(df):
            series = df[col]
            # Για τις ημερομηνίες, το str(cell.value) μπορεί να είναι μακρύ, οπότε ορίζουμε ένα σταθερό πλάτος.
            if 'Ημερομηνία' in str(series.name):
                max_len = 15
            else:
                 max_len = max((
                    series.astype(str).map(len).max(),
                    len(str(series.name))
                    )) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len

    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="orders_export.xlsx"'

    return response
@login_required
def export_payments_to_excel(request):
    # Παίρνουμε το ίδιο queryset και τους ίδιους κανόνες φιλτραρίσματος από τη view all_payments_list_view
    payments_qs = Payment.objects.select_related(
        'customer', 'order', 'recorded_by', 'cancelled_by'
    ).all().order_by('-payment_date', '-receipt_number')

    # Λήψη τιμών φίλτρων από το GET request
    query_text = request.GET.get('q', '').strip()
    customer_filter_id = request.GET.get('customer_filter', None)
    status_filter = request.GET.get('status_filter', 'all')
    method_filter = request.GET.get('method_filter', 'all')
    date_from_str = request.GET.get('date_from', None)
    date_to_str = request.GET.get('date_to', None)
    value_date_from_str = request.GET.get('value_date_from', None)
    value_date_to_str = request.GET.get('value_date_to', None)

    # Εφαρμογή φίλτρων
    if query_text:
        payments_qs = payments_qs.filter(
            Q(receipt_number__icontains=query_text) |
            Q(reference_number__icontains=query_text) |
            Q(notes__icontains=query_text)
        )
    if customer_filter_id:
        try:
            payments_qs = payments_qs.filter(customer_id=int(customer_filter_id))
        except (ValueError, TypeError): pass
    if status_filter != 'all':
        payments_qs = payments_qs.filter(status=status_filter)
    if method_filter != 'all':
        payments_qs = payments_qs.filter(payment_method=method_filter)
    if date_from_str:
        try:
            payments_qs = payments_qs.filter(payment_date__gte=datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date())
        except ValueError: pass
    if date_to_str:
        try:
            payments_qs = payments_qs.filter(payment_date__lte=datetime.datetime.strptime(date_to_str, '%Y-%m-%d').date())
        except ValueError: pass
    if value_date_from_str:
        try:
            payments_qs = payments_qs.filter(value_date__gte=datetime.datetime.strptime(value_date_from_str, '%Y-%m-%d').date())
        except ValueError: pass
    if value_date_to_str:
        try:
            payments_qs = payments_qs.filter(value_date__lte=datetime.datetime.strptime(value_date_to_str, '%Y-%m-%d').date())
        except ValueError: pass

    # Προετοιμασία δεδομένων για το DataFrame
    data_for_export = []
    for payment in payments_qs:
        data_for_export.append({
            'Αρ. Απόδειξης Συστήματος': payment.receipt_number,
            'Ημερομηνία Πληρωμής': payment.payment_date,
            'Πελάτης': str(payment.customer) if payment.customer else '',
            'Ποσό (€)': payment.amount_paid,
            'Τρόπος Πληρωμής': payment.get_payment_method_display(),
            'Κατάσταση': payment.get_status_display(),
            'Εξωτερικός Αρ. Αναφοράς': payment.reference_number,
            'Ημερομηνία Λήξης/Value': payment.value_date,
            'Σχετ. Παραγγελία': payment.order.order_number if payment.order else '',
            'Σημειώσεις': payment.notes,
            'Λόγος Ακύρωσης': payment.cancellation_reason,
            'Καταχωρήθηκε από': payment.recorded_by.username if payment.recorded_by else '',
            'Ακυρώθηκε από': payment.cancelled_by.username if payment.cancelled_by else '',
            'Ημερομηνία Ακύρωσης': payment.cancelled_at.strftime('%Y-%m-%d %H:%M:%S') if payment.cancelled_at else '',
        })

    if not data_for_export:
        messages.warning(request, "Δεν βρέθηκαν πληρωμές για εξαγωγή με βάση τα επιλεγμένα κριτήρια.")
        return redirect(request.META.get('HTTP_REFERER', 'all_payments_list'))

    df = pd.DataFrame(data_for_export)

    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Πληρωμές')
        worksheet = writer.sheets['Πληρωμές']
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(df):
            series = df[col]
            # Για τις ημερομηνίες, το str(cell.value) μπορεί να είναι μακρύ, οπότε ορίζουμε ένα σταθερό πλάτος.
            if 'Ημερομηνία' in str(series.name):
                max_len = 15
            else:
                max_len = max((
                    series.astype(str).map(len).max(),
                    len(str(series.name))
                    )) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len

    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="payments_export.xlsx"'

    return response
@login_required
def export_stock_receipts_to_excel(request):
    # Παίρνουμε το ίδιο queryset και τους ίδιους κανόνες φιλτραρίσματος από τη view stock_receipt_list
    receipts_qs = StockReceipt.objects.select_related('product', 'user_who_recorded').all().order_by('-date_received')

    query = request.GET.get('q', '').strip()
    product_filter_id = request.GET.get('product_filter', None)
    date_from_str = request.GET.get('date_from', None)
    date_to_str = request.GET.get('date_to', None)

    # Εφαρμογή φίλτρων
    if product_filter_id:
        try:
            receipts_qs = receipts_qs.filter(product_id=int(product_filter_id))
        except (ValueError, TypeError): pass
    if date_from_str:
        try:
            receipts_qs = receipts_qs.filter(date_received__date__gte=datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date())
        except ValueError: pass
    if date_to_str:
        try:
            receipts_qs = receipts_qs.filter(date_received__date__lte=datetime.datetime.strptime(date_to_str, '%Y-%m-%d').date())
        except ValueError: pass
    if query:
        normalized_query = normalize_text(query)
        receipts_qs = receipts_qs.filter(
            Q(product__name_normalized__icontains=normalized_query) |
            Q(notes__icontains=query) |
            Q(user_who_recorded__username__icontains=query)
        ).distinct()

    # Προετοιμασία δεδομένων για το DataFrame
    data_for_export = []
    for receipt in receipts_qs:
        data_for_export.append({
            'Προϊόν': receipt.product.name if receipt.product else '',
            'Κωδικός Προϊόντος': receipt.product.code if receipt.product else '',
            'Ποσότητα Εισαγωγής': receipt.quantity_added,
            'Μονάδα Μέτρησης': receipt.product.get_unit_of_measurement_display() if receipt.product else '',
            'Ημερομηνία & Ώρα Παραλαβής': receipt.date_received.strftime('%Y-%m-%d %H:%M:%S') if receipt.date_received else '',
            'Χρήστης Καταχώρησης': receipt.user_who_recorded.username if receipt.user_who_recorded else '',
            'Σημειώσεις': receipt.notes,
        })

    if not data_for_export:
        messages.warning(request, "Δεν βρέθηκαν παραλαβές για εξαγωγή με βάση τα επιλεγμένα κριτήρια.")
        return redirect(request.META.get('HTTP_REFERER', 'stock_receipt_list'))

    df = pd.DataFrame(data_for_export)
    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Παραλαβές Αποθέματος')
        worksheet = writer.sheets['Παραλαβές Αποθέματος']
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(df):
            series = df[col]
            if 'Ημερομηνία' in str(series.name):
                max_len = 20
            else:
                max_len = max((series.astype(str).map(len).max(), len(str(series.name)))) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len

    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stock_receipts_export.xlsx"'
    return response
@login_required
def stock_overview_list(request):
    products_qs = Product.objects.all().order_by('name')

    # Φιλτράρισμα
    status_filter = request.GET.get('status_filter', 'all')
    query_text = request.GET.get('q', '').strip()

    if query_text:
        normalized_query = normalize_text(query_text)
        products_qs = products_qs.filter(
            Q(name_normalized__icontains=normalized_query) |
            Q(code__icontains=query_text) |
            Q(barcode__icontains=query_text)
        ).distinct()
    
    if status_filter == 'low':
        # Προϊόντα όπου το απόθεμα είναι μικρότερο ή ίσο του ορίου ασφαλείας
        products_qs = products_qs.filter(stock_quantity__lte=models.F('min_stock_level'))
    elif status_filter == 'out_of_stock':
        # Προϊόντα όπου το απόθεμα είναι μηδέν ή λιγότερο
        products_qs = products_qs.filter(stock_quantity__lte=0)
    elif status_filter == 'sufficient':
        # Προϊόντα όπου το απόθεμα είναι πάνω από το όριο ασφαλείας
        products_qs = products_qs.filter(stock_quantity__gt=models.F('min_stock_level'))
    
    
    # Σελιδοποίηση
    paginator = Paginator(products_qs, 30) # Π.χ. 30 προϊόντα ανά σελίδα
    page_number = request.GET.get('page')
    try:
        products_page = paginator.page(page_number)
    except PageNotAnInteger:
        products_page = paginator.page(1)
    except EmptyPage:
        products_page = paginator.page(paginator.num_pages)

    context = {
        'products': products_page,
        'status_filter': status_filter,
        'query_text': query_text,
        'title': 'Επισκόπηση Αποθεμάτων'
    }
    return render(request, 'core/stock_overview_list.html', context)
@login_required
def export_stock_overview_excel(request):
    # Παίρνουμε το ίδιο queryset και τους ίδιους κανόνες φιλτραρίσματος από τη view stock_overview_list
    products_qs = Product.objects.all().order_by('name')

    status_filter = request.GET.get('status_filter', 'all')
    query_text = request.GET.get('q', '').strip()

    if query_text:
        normalized_user_query = normalize_text(query_text)
        products_qs = products_qs.filter(
            Q(name_normalized__icontains=normalized_user_query) |
            Q(code__icontains=query_text) |
            Q(barcode__icontains=query_text)
        ).distinct()
    
    if status_filter == 'low':
        products_qs = products_qs.filter(stock_quantity__lte=models.F('min_stock_level'))
    elif status_filter == 'out_of_stock':
        products_qs = products_qs.filter(stock_quantity__lte=0)
    elif status_filter == 'sufficient':
        products_qs = products_qs.filter(stock_quantity__gt=models.F('min_stock_level'))
    
    # Προετοιμασία δεδομένων για το DataFrame
    data_for_export = []
    for product in products_qs:
        # Υπολογισμός της κατάστασης αποθέματος
        stock_status = 'Επαρκές'
        if product.stock_quantity <= 0:
            stock_status = 'Εκτός Αποθέματος'
        elif product.stock_quantity <= product.min_stock_level:
            stock_status = 'Χαμηλό Απόθεμα'

        data_for_export.append({
            'Όνομα Προϊόντος': product.name,
            'Κωδικός': product.code,
            'Τρέχον Απόθεμα': product.stock_quantity,
            'Ελάχιστο Όριο': product.min_stock_level,
            'Μονάδα Μέτρησης': product.get_unit_of_measurement_display(),
            'Κατάσταση Αποθέματος': stock_status,
        })
    
    if not data_for_export:
        messages.warning(request, "Δεν βρέθηκαν προϊόντα για εξαγωγή με βάση τα επιλεγμένα κριτήρια.")
        return redirect(request.META.get('HTTP_REFERER', 'stock_overview_list'))

    df = pd.DataFrame(data_for_export)

    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Επισκόπηση Αποθεμάτων')
        worksheet = writer.sheets['Επισκόπηση Αποθεμάτων']
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(df):
            series = df[col]
            max_len = max((
                series.astype(str).map(len).max(),
                len(str(series.name))
                )) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len
    
    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stock_overview_export.xlsx"'

    return response
@login_required
@require_POST
def order_create_invoice_view(request, order_pk):
    order = get_object_or_404(Order.objects.select_related('customer__parent'), pk=order_pk)
    
    billing_customer = order.customer
    if order.customer.is_branch:
        billing_customer = order.customer.parent
        if not billing_customer:
            messages.error(request, f"Το υποκατάστημα '{order.customer}' δεν έχει δηλωμένο κεντρικό κατάστημα και δεν μπορεί να τιμολογηθεί.")
            return redirect('order_detail', pk=order.pk)

    if Invoice.objects.filter(order=order, customer=billing_customer).exists():
        invoice = Invoice.objects.get(order=order, customer=billing_customer)
        messages.warning(request, f"Έχει ήδη εκδοθεί το τιμολόγιο {invoice.invoice_number} για αυτή την παραγγελία.")
        return redirect('invoice_detail', pk=invoice.pk)
        
    if order.status != Order.STATUS_COMPLETED:
        messages.error(request, "Μόνο οι ολοκληρωμένες παραγγελίες μπορούν να τιμολογηθούν.")
        return redirect('order_detail', pk=order.pk)

    try:
        with transaction.atomic():
            invoice = Invoice.objects.create(
                customer=billing_customer,
                order=order,
                issue_date=timezone.now().date(),
                status=Invoice.STATUS_ISSUED
            )

            # Αυτόματη αντιγραφή στοιχείων διακίνησης
            source = order.delivery_notes.last() or order
            invoice.purpose = source.purpose
            invoice.carrier = source.carrier
            invoice.license_plate = source.license_plate
            invoice.shipping_name = source.shipping_name
            invoice.shipping_address = source.shipping_address
            invoice.shipping_city = source.shipping_city
            invoice.shipping_postal_code = source.shipping_postal_code
            if hasattr(source, 'shipping_vat_number'):
                invoice.shipping_vat_number = source.shipping_vat_number
            
            # --- ΕΔΩ ΕΙΝΑΙ Η ΔΙΟΡΘΩΣΗ ---
            # Δημιουργία των ειδών του τιμολογίου ΜΕ ΥΠΟΛΟΓΙΣΜΟ ΤΙΜΩΝ
            for order_item in order.items.all():
                
                # Υπολογίζουμε την καθαρή αξία της γραμμής (μετά την έκπτωση είδους)
                price_after_discount = order_item.unit_price * (Decimal('1') - (order_item.discount_percentage / Decimal('100')))
                line_subtotal = order_item.quantity * price_after_discount
                
                # Υπολογίζουμε το ΦΠΑ της γραμμής
                line_vat_amount = line_subtotal * (order_item.vat_percentage / Decimal('100'))

                InvoiceItem.objects.create(
                    invoice=invoice,
                    product=order_item.product,
                    description=order_item.product.name if order_item.product else "Είδος χωρίς προϊόν",
                    quantity=order_item.quantity,
                    unit_price=order_item.unit_price,
                    is_gift=order_item.is_gift,
                    discount_percentage=order_item.discount_percentage,
                    vat_percentage=order_item.vat_percentage,
                    # Περνάμε τις υπολογισμένες τιμές
                    total_price=line_subtotal,
                    vat_amount=line_vat_amount
                )
            # --- ΤΕΛΟΣ ΔΙΟΡΘΩΣΗΣ ---
            
            # Τώρα ο υπολογισμός των συνόλων θα δουλέψει σωστά
            invoice.calculate_totals()
            invoice.save() # Αποθηκεύουμε το τιμολόγιο με τα σωστά σύνολα
            
            # Ενημέρωση υπολοίπου πελάτη
            billing_customer.balance += invoice.total_amount
            billing_customer.save(update_fields=['balance', 'updated_at'])
            
            messages.success(request, f"Το τιμολόγιο {invoice.invoice_number} δημιουργήθηκε για τον πελάτη {billing_customer}.")
            return redirect('invoice_detail', pk=invoice.pk)

    except Exception as e:
        messages.error(request, f"Προέκυψε ένα μη αναμενόμενο σφάλμα: {e}")
        return redirect('order_detail', pk=order.pk)
def invoice_list_view(request):
    user = request.user
    
    # Ξεκινάμε με το βασικό queryset
    invoices_qs = Invoice.objects.select_related('customer', 'order')

    # Εφαρμόζουμε τους κανόνες δικαιωμάτων
    if not user.is_superuser:
        if user.groups.filter(name='Πωλητές').exists():
            try:
                invoices_qs = invoices_qs.filter(customer__sales_rep=user.salesrepresentative)
            except SalesRepresentative.DoesNotExist:
                invoices_qs = Invoice.objects.none()
        # Οι ομάδες 'Λογιστήριο' και 'Αποθήκη' μπορούν να βλέπουν όλα τα τιμολόγια.

    # Η υπόλοιπη λογική των φίλτρων παραμένει ίδια
    query_text = request.GET.get('q', '').strip()
    customer_filter_id = request.GET.get('customer_filter', None)
    status_filter = request.GET.get('status_filter', 'all')
    date_from_str = request.GET.get('date_from', None)
    date_to_str = request.GET.get('date_to', None)
    due_filter = request.GET.get('due', None)

    if query_text:
        invoices_qs = invoices_qs.filter(invoice_number__icontains=query_text)
    if customer_filter_id:
        invoices_qs = invoices_qs.filter(customer_id=customer_filter_id)
    if status_filter != 'all':
        invoices_qs = invoices_qs.filter(status=status_filter)
    if date_from_str:
        try:
            invoices_qs = invoices_qs.filter(issue_date__gte=date_from_str)
        except: pass
    if date_to_str:
        try:
            invoices_qs = invoices_qs.filter(issue_date__lte=date_to_str)
        except: pass
            
    today = timezone.now().date()
    if due_filter == 'soon':
        due_date_threshold = today + timedelta(days=7)
        invoices_qs = invoices_qs.filter(
            status=Invoice.STATUS_ISSUED,
            due_date__gte=today,
            due_date__lte=due_date_threshold
        ).order_by('due_date')
    elif due_filter == 'overdue':
        invoices_qs = invoices_qs.filter(
            status=Invoice.STATUS_ISSUED,
            due_date__lt=today
        ).order_by('due_date')
    else:
        invoices_qs = invoices_qs.order_by('-issue_date', '-pk')

    paginator = Paginator(invoices_qs, 20)
    page_number = request.GET.get('page')
    invoices_page = paginator.get_page(page_number)

    customers_for_filter = Customer.objects.all().order_by('last_name', 'first_name')
    invoice_status_choices = Invoice.INVOICE_STATUS_CHOICES

    context = {
        'invoices': invoices_page,
        'title': 'Λίστα Τιμολογίων',
        'query_text': query_text,
        'customer_filter_id': int(customer_filter_id) if customer_filter_id else None,
        'status_filter': status_filter,
        'date_from_value': date_from_str,
        'date_to_value': date_to_str,
        'customers_for_filter': customers_for_filter,
        'invoice_status_choices': invoice_status_choices,
    }
    return render(request, 'core/invoice_list.html', context)
@login_required
def invoice_detail_view(request, pk):
    # Χρησιμοποιούμε select_related για να πάρουμε και τα στοιχεία του πελάτη με ένα μόνο query
    invoice = get_object_or_404(Invoice.objects.select_related('customer'), pk=pk)
    context = {
        'invoice': invoice,
        'title': f'Τιμολόγιο {invoice.invoice_number}',
        'today': timezone.now().date()
    }
    return render(request, 'core/invoice_detail.html', context)
@login_required
@require_POST
def invoice_mark_as_paid_view(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    
    if invoice.status == Invoice.STATUS_ISSUED:
        with transaction.atomic():
            # 1. Αλλάζουμε την κατάσταση του τιμολογίου
            invoice.status = Invoice.STATUS_PAID
            invoice.save(update_fields=['status', 'updated_at'])

            # 2. Δημιουργούμε μια αυτόματη εγγραφή Πληρωμής
            auto_payment = Payment.objects.create(
                customer=invoice.customer,
                amount_paid=invoice.total_amount,
                payment_date=timezone.now().date(),
                payment_method='other', # 'other' ή μια άλλη default τιμή
                notes=f"Αυτόματη εξόφληση βάσει του τιμολογίου {invoice.invoice_number}",
                recorded_by=request.user,
                status=Payment.STATUS_ACTIVE
            )
            
            # 3. Συνδέουμε την νέα πληρωμή με το τιμολόγιο
            auto_payment.invoices.add(invoice)

            # 4. Η .save() της Πληρωμής έχει ήδη μειώσει το balance, οπότε δεν χρειάζεται να το ξανακάνουμε.
            # Η λογική είναι συνεπής.
            
        messages.success(request, f"Το τιμολόγιο {invoice.invoice_number} σημειώθηκε ως εξοφλημένο και δημιουργήθηκε αυτόματη πληρωμή.")
    else:
        messages.warning(request, f"Το τιμολόγιο είναι σε κατάσταση '{invoice.get_status_display()}' και δεν μπορεί να αλλάξει.")
        
    return redirect('invoice_detail', pk=invoice.pk)

@login_required
@require_POST
def invoice_mark_as_issued_view(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    if invoice.status == Invoice.STATUS_DRAFT:
        invoice.status = Invoice.STATUS_ISSUED
        invoice.save(update_fields=['status', 'updated_at'])
        messages.success(request, f"Το τιμολόγιο {invoice.invoice_number} εκδόθηκε επιτυχώς.")
    else:
        messages.warning(request, "Αυτό το τιμολόγιο έχει ήδη εκδοθεί.")
    return redirect('invoice_detail', pk=invoice.pk)
@login_required
@require_POST
def invoice_cancel_view(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)

    # Έλεγχος αν επιτρέπεται η ακύρωση
    if invoice.issue_date != timezone.now().date() or invoice.status != Invoice.STATUS_ISSUED:
        messages.error(request, "Το τιμολόγιο μπορεί να ακυρωθεί μόνο την ίδια μέρα που εκδόθηκε και εφόσον είναι σε κατάσταση 'Εκδόθηκε'.")
        return redirect('invoice_detail', pk=invoice.pk)

    try:
        with transaction.atomic():
            # 1. Ακύρωση του Τιμολογίου
            invoice.status = Invoice.STATUS_CANCELLED
            invoice.save(update_fields=['status'])

            # 2. Αντιλογισμός του ποσού από το υπόλοιπο του πελάτη
            invoice.customer.balance -= invoice.total_amount
            invoice.customer.save(update_fields=['balance'])

            # --- ΔΙΟΡΘΩΜΕΝΗ ΛΟΓΙΚΗ ΑΚΥΡΩΣΗΣ ΑΛΥΣΙΔΑΣ ---
            related_order = None
            
            # 3. Έλεγχος και ακύρωση του Δελτίου Αποστολής (ΑΝ ΥΠΑΡΧΕΙ)
            if invoice.delivery_note:
                delivery_note_to_cancel = invoice.delivery_note
                delivery_note_to_cancel.status = DeliveryNote.Status.CANCELLED
                delivery_note_to_cancel.save(update_fields=['status'])
                
                # Η παραγγελία είναι συνδεδεμένη μέσω του Δ.Α.
                if delivery_note_to_cancel.order:
                    related_order = delivery_note_to_cancel.order

            # Αν δεν υπήρχε Δ.Α., η παραγγελία είναι συνδεδεμένη απευθείας με το τιμολόγιο
            elif invoice.order:
                related_order = invoice.order

            # 4. Ακύρωση της αρχικής Παραγγελίας (που θα επιστρέψει το απόθεμα)
            if related_order:
                related_order.status = Order.STATUS_CANCELLED
                related_order.save(update_fields=['status'])
            # --- ΤΕΛΟΣ ΔΙΟΡΘΩΜΕΝΗΣ ΛΟΓΙΚΗΣ ---

        messages.success(request, f"Το τιμολόγιο {invoice.invoice_number} ακυρώθηκε επιτυχώς. Το σχετικό Δελτίο Αποστολής και η Παραγγελία ενημερώθηκαν.")
    except Exception as e:
        messages.error(request, f"Προέκυψε ένα σφάλμα κατά την ακύρωση: {e}")

    return redirect('invoice_detail', pk=invoice.pk)
@login_required
def export_invoices_to_excel(request):
    # Εδώ θα μπορούσαμε να εφαρμόσουμε τα ίδια φίλτρα που έχει η λίστας μας
    invoices_qs = Invoice.objects.select_related('customer', 'order').all()

    data_for_export = []
    for invoice in invoices_qs:
        data_for_export.append({
            'Αρ. Τιμολογίου': invoice.invoice_number,
            'Ημερομηνία Έκδοσης': invoice.issue_date,
            'Πελάτης': str(invoice.customer),
            'Κατάσταση': invoice.get_status_display(),
            'Υποσύνολο (€)': invoice.subtotal,
            'ΦΠΑ (€)': invoice.vat_amount,
            'Τελικό Σύνολο (€)': invoice.total_amount,
            'Σχετ. Παραγγελία': invoice.order.order_number if invoice.order else '',
        })

    if not data_for_export:
        messages.warning(request, "Δεν βρέθηκαν τιμολόγια για εξαγωγή.")
        return redirect('invoice_list')

    df = pd.DataFrame(data_for_export)
    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Τιμολόγια')
        worksheet = writer.sheets['Τιμολόγια']
        for idx, col in enumerate(df):
            series = df[col]
            max_len = max((series.astype(str).map(len).max(), len(str(series.name)))) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len

    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="invoices_export.xlsx"'
    return response
@login_required
def invoice_pdf_view(request, pk):
    if not WEASYPRINT_AVAILABLE:
        raise Http404("Η βιβλιοθήκη WeasyPrint λείπει. Η δημιουργία PDF δεν είναι δυνατή.")

    invoice = get_object_or_404(Invoice.objects.select_related('customer'), pk=pk)

    # --- ΠΡΟΣΘΗΚΗ ΕΛΕΓΧΟΥ ΑΣΦΑΛΕΙΑΣ ---
    if invoice.status == Invoice.STATUS_CANCELLED:
        messages.error(request, "Αυτό το τιμολόγιο είναι ακυρωμένο. Μπορείτε να εκτυπώσετε μόνο το ακυρωτικό του σημείωμα.")
        return redirect('invoice_detail', pk=pk)
    # --- ΤΕΛΟΣ ΕΛΕΓΧΟΥ ---

    company_info = getattr(settings, 'COMPANY_INFO', {})
    context = {
        'invoice': invoice,
        'company_info': company_info,
    }

    html_string = render_to_string('core/invoice_pdf.html', context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    response = HttpResponse(html.write_pdf(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="invoice_{invoice.invoice_number}.pdf"'
    return response
@staff_member_required # Μόνο οι διαχειριστές μπορούν να δουν τη λίστα
def sales_rep_list_view(request):
    sales_reps = SalesRepresentative.objects.select_related('user').all()
    context = {
        'sales_reps': sales_reps,
        'title': 'Διαχείριση Πωλητών/Αντιπροσώπων'
    }
    return render(request, 'core/sales_rep_list.html', context)

@staff_member_required
def sales_rep_create_view(request):
    if request.method == 'POST':
        form = SalesRepresentativeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ο πωλητής/αντιπρόσωπος δημιουργήθηκε επιτυχώς.')
            return redirect('sales_rep_list')
    else:
        form = SalesRepresentativeForm()
    
    context = {
        'form': form,
        'title': 'Δημιουργία Νέου Πωλητή/Αντιπροσώπου'
    }
    return render(request, 'core/sales_rep_form.html', context)

@staff_member_required
def sales_rep_edit_view(request, pk):
    sales_rep = get_object_or_404(SalesRepresentative, pk=pk)
    if request.method == 'POST':
        form = SalesRepresentativeForm(request.POST, instance=sales_rep)
        if form.is_valid():
            form.save()
            messages.success(request, 'Οι αλλαγές αποθηκεύτηκαν επιτυχώς.')
            return redirect('sales_rep_list')
    else:
        form = SalesRepresentativeForm(instance=sales_rep)

    context = {
        'form': form,
        'sales_rep': sales_rep,
        'title': f'Επεξεργασία: {sales_rep}'
    }
    return render(request, 'core/sales_rep_form.html', context)

@staff_member_required
@require_POST
def sales_rep_delete_view(request, pk):
    sales_rep = get_object_or_404(SalesRepresentative, pk=pk)
    rep_name = str(sales_rep)
    sales_rep.delete()
    messages.success(request, f'Ο πωλητής/αντιπρόσωπος "{rep_name}" διαγράφηκε επιτυχώς.')
    return redirect('sales_rep_list')
@staff_member_required
def commission_report_view(request):
    # Χειρισμός της υποβολής της φόρμας για την πληρωμή προμηθειών
    if request.method == 'POST':
        commission_ids = request.POST.getlist('commission_ids')
        if not commission_ids:
            messages.warning(request, 'Δεν επιλέξατε καμία προμήθεια για εξόφληση.')
            return redirect('commission_report')

        commissions_to_pay = Commission.objects.filter(id__in=commission_ids, status=Commission.Status.UNPAID)
        updated_count = commissions_to_pay.update(
            status=Commission.Status.PAID, 
            paid_date=timezone.now().date()
        )
        
        if updated_count > 0:
            messages.success(request, f'{updated_count} προμήθειες σημειώθηκαν ως εξοφλημένες.')
        else:
            messages.info(request, 'Οι επιλεγμένες προμήθειες ήταν ήδη εξοφλημένες.')
        
        return redirect('commission_report')

    # Λογική για την εμφάνιση της σελίδας (GET request)
    commissions_qs = Commission.objects.select_related(
        'sales_rep__user', 'invoice__customer'
    ).all()

    # Φιλτράρισμα
    rep_filter_id = request.GET.get('sales_rep', None)
    status_filter = request.GET.get('status', 'UNPAID') # Προεπιλογή να δείχνει τις ανεξόφλητες

    if rep_filter_id:
        commissions_qs = commissions_qs.filter(sales_rep_id=rep_filter_id)
    if status_filter:
        commissions_qs = commissions_qs.filter(status=status_filter)

    # Υπολογισμός συνόλου για τις ανεξόφλητες προμήθειες (με βάση τα φίλτρα)
    total_unpaid = commissions_qs.filter(status=Commission.Status.UNPAID).aggregate(
        total=Sum('calculated_amount')
    )['total'] or 0.00

    all_reps = SalesRepresentative.objects.select_related('user').all()

    context = {
        'title': 'Αναφορά & Διαχείριση Προμηθειών',
        'commissions': commissions_qs,
        'total_unpaid': total_unpaid,
        'all_reps': all_reps,
        'rep_filter_id': rep_filter_id,
        'status_filter': status_filter,
        'commission_statuses': Commission.Status.choices
    }
    return render(request, 'core/commission_report.html', context)
@staff_member_required
def create_credit_note_from_invoice(request, invoice_pk):
    original_invoice = get_object_or_404(Invoice, pk=invoice_pk)
    
    if original_invoice.status not in [Invoice.STATUS_ISSUED, Invoice.STATUS_PAID]:
        messages.error(request, f"Δεν μπορείτε να εκδώσετε πιστωτικό για τιμολόγιο που είναι σε κατάσταση '{original_invoice.get_status_display()}'.")
        return redirect('invoice_detail', pk=original_invoice.pk)

    initial_formset_data = [{
        'product': item.product,
        'description': item.description,
        'unit_price': item.unit_price,
        'vat_percentage': item.vat_percentage,
        'quantity': 0, # Αρχικοποιούμε την ποσότητα επιστροφής σε 0
        'original_quantity': item.quantity,
        'is_gift': item.is_gift,
    } for item in original_invoice.items.all()]

    if request.method == 'POST':
        form = CreditNoteForm(request.POST)
        formset = CreditNoteItemFormSet(request.POST, initial=initial_formset_data)

        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    credit_note = form.save(commit=False)
                    credit_note.customer = original_invoice.customer
                    credit_note.original_invoice = original_invoice
                    credit_note.save()

                    total_subtotal = Decimal('0.00')
                    total_vat = Decimal('0.00')
                    items_returned = False

                    for item_form in formset:
                        cleaned_data = item_form.cleaned_data
                        returned_quantity = cleaned_data.get('quantity')
                        
                        if returned_quantity and returned_quantity > 0:
                            items_returned = True
                            
                            # --- ΕΔΩ ΕΙΝΑΙ Η ΔΙΟΡΘΩΣΗ ---
                            # 1. Υπολογίζουμε τις τιμές ΠΡΙΝ τη δημιουργία
                            line_total = returned_quantity * cleaned_data['unit_price']
                            line_vat = line_total * (cleaned_data['vat_percentage'] / Decimal('100'))

                            # 2. Περνάμε τις υπολογισμένες τιμές απευθείας στο create()
                            cn_item = CreditNoteItem.objects.create(
                                credit_note=credit_note,
                                product=cleaned_data['product'],
                                description=cleaned_data['description'],
                                quantity=returned_quantity,
                                unit_price=cleaned_data['unit_price'],
                                vat_percentage=cleaned_data['vat_percentage'],
                                total_price=line_total,   # <--- Προσθήκη
                                vat_amount=line_vat       # <--- Προσθήκη
                            )
                            
                            # Το stock θα επιστραφεί από το signal του Product που έχεις ήδη (αν το έχεις ορίσει)
                            # ή θα έπρεπε να προστεθεί εδώ η λογική:
                            if cn_item.product:
                                cn_item.product.stock_quantity += cn_item.quantity
                                cn_item.product.save(update_fields=['stock_quantity'])
                            
                            total_subtotal += line_total
                            total_vat += line_vat
                    
                    if not items_returned:
                        messages.warning(request, "Δεν επιλέξατε ποσότητα επιστροφής για κανένα είδος.")
                        # Χρησιμοποιούμε transaction.set_rollback(True) για να ακυρώσουμε τη συναλλαγή
                        transaction.set_rollback(True)
                    else:
                        credit_note.subtotal = total_subtotal
                        credit_note.vat_amount = total_vat
                        credit_note.total_amount = total_subtotal + total_vat
                        credit_note.status = CreditNote.Status.ISSUED
                        credit_note.save()
                        
                        customer_to_update = Customer.objects.select_for_update().get(pk=credit_note.customer.pk)
                        customer_to_update.balance -= credit_note.total_amount
                        customer_to_update.save(update_fields=['balance'])
                        
                        original_invoice.status = Invoice.STATUS_CREDITED
                        original_invoice.save(update_fields=['status'])

                        messages.success(request, f"Το πιστωτικό τιμολόγιο {credit_note.credit_note_number} δημιουργήθηκε επιτυχώς.")
                        return redirect('credit_note_detail', pk=credit_note.pk)

            except Exception as e:
                 messages.error(request, f"Προέκυψε ένα άγνωστο σφάλμα: {e}")

        else:
            messages.error(request, "Η φόρμα περιέχει σφάλματα. Παρακαλώ διορθώστε τα παρακάτω.")

    else: # GET Request
        form = CreditNoteForm(initial={'reason': f'Επιστροφή/Ακύρωση ειδών από το Τιμολόγιο {original_invoice.invoice_number}'})
        formset = CreditNoteItemFormSet(initial=initial_formset_data)

    context = {
        'title': f"Δημιουργία Πιστωτικού για το Τιμολόγιο {original_invoice.invoice_number}",
        'form': form,
        'formset': formset,
        'original_invoice': original_invoice
    }
    return render(request, 'core/credit_note_form.html', context)
@login_required
def user_profile_edit_view(request):
    # Με το get_or_create, διασφαλίζουμε ότι θα υπάρχει προφίλ ακόμα και για παλιούς χρήστες
    # που δημιουργήθηκαν πριν το signal.
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, 'Οι ρυθμίσεις εμφάνισης αποθηκεύτηκαν επιτυχώς!')
            return redirect('user_profile_settings')
    else:
        form = UserProfileForm(instance=profile)
    
    context = {
        'title': 'Ρυθμίσεις Εμφάνισης',
        'form': form
    }
    return render(request, 'core/user_profile_form.html', context)
@login_required
def credit_note_detail_view(request, pk):
    credit_note = get_object_or_404(CreditNote.objects.select_related('customer', 'original_invoice'), pk=pk)
    context = {
        'title': f'Πιστωτικό Τιμολόγιο {credit_note.credit_note_number}',
        'credit_note': credit_note
    }
    return render(request, 'core/credit_note_detail.html', context)

@login_required
def credit_note_pdf_view(request, pk):
    if not WEASYPRINT_AVAILABLE:
        raise Http404("Η βιβλιοθήκη WeasyPrint λείπει. Η δημιουργία PDF δεν είναι δυνατή.")

    credit_note = get_object_or_404(CreditNote.objects.select_related('customer'), pk=pk)
    company_info = getattr(settings, 'COMPANY_INFO', {})

    context = {
        'credit_note': credit_note,
        'company_info': company_info,
    }

    html_string = render_to_string('core/credit_note_pdf.html', context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))

    response = HttpResponse(html.write_pdf(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="credit_note_{credit_note.credit_note_number}.pdf"'
    return response
@login_required
def credit_note_list_view(request):
    user = request.user
    credit_notes_qs = CreditNote.objects.select_related('customer', 'original_invoice').all()

    # Εφαρμόζουμε τα δικαιώματα πρόσβασης
    if not user.is_superuser:
        if user.groups.filter(name='Πωλητές').exists():
            try:
                credit_notes_qs = credit_notes_qs.filter(customer__sales_rep=user.salesrepresentative)
            except SalesRepresentative.DoesNotExist:
                credit_notes_qs = CreditNote.objects.none()

    paginator = Paginator(credit_notes_qs.order_by('-issue_date', '-pk'), 20)
    page_number = request.GET.get('page')
    credit_notes_page = paginator.get_page(page_number)

    context = {
        'title': 'Λίστα Πιστωτικών Τιμολογίων',
        'credit_notes': credit_notes_page
    }
    return render(request, 'core/credit_note_list.html', context)
def my_profile_edit_view(request):
    if request.method == 'POST':
        form = MyProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Το προφίλ σας ενημερώθηκε επιτυχώς!')
            return redirect('my_profile')
    else:
        form = MyProfileForm(instance=request.user)

    context = {
        'title': 'Το Προφίλ μου',
        'form': form
    }
    return render(request, 'core/my_profile_form.html', context)
@staff_member_required
def retail_pos_view(request):
    if request.method == 'POST':
        # Πλέον έχουμε δύο φόρμες για έλεγχο
        form = RetailReceiptForm(request.POST)
        formset = RetailReceiptItemFormSet(request.POST, prefix='items') # Προσθέτουμε το σωστό prefix

        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    # --- ΝΕΑ ΛΟΓΙΚΗ ΕΠΙΛΟΓΗΣ ΠΕΛΑΤΗ ---
                    selected_customer = form.cleaned_data.get('customer')

                    if selected_customer:
                        receipt_customer = selected_customer
                    else:
                        # Fallback στον default πελάτη λιανικής
                        try:
                            receipt_customer = Customer.objects.get(code='RETAIL')
                        except Customer.DoesNotExist:
                            messages.error(request, 'Δεν βρέθηκε ο default "Πελάτης Λιανικής". Παρακαλώ δημιουργήστε τον με κωδικό "RETAIL".')
                            raise Exception("Retail customer not found")

                    receipt = RetailReceipt.objects.create(customer=receipt_customer)

                    total_subtotal = Decimal('0.00')
                    total_vat = Decimal('0.00')
                    items_added = False

                    for item_form in formset:
                        if item_form.cleaned_data and item_form.cleaned_data.get('product'):
                            items_added = True
                            product = item_form.cleaned_data['product']
                            quantity = item_form.cleaned_data['quantity']
                            discount = item_form.cleaned_data.get('discount_percentage', 0)

                            if quantity > product.stock_quantity:
                                messages.error(request, f"Το απόθεμα για το προϊόν '{product.name}' δεν επαρκεί (Διαθέσιμο: {product.stock_quantity}).")
                                raise Exception("Insufficient stock")

                            base_price = product.price # Αυτή είναι η καθαρή τιμή
                            price_after_discount = base_price * (Decimal('1') - (discount / Decimal('100')))

                            # Υπολογισμοί ανά γραμμή
                            line_subtotal = quantity * price_after_discount
                            line_vat = line_subtotal * (product.vat_percentage / Decimal('100'))
                            line_final_price = line_subtotal + line_vat

                            RetailReceiptItem.objects.create(
                                receipt=receipt, product=product, description=product.name,
                                quantity=quantity, unit_price=base_price, discount_percentage=discount,
                                final_price=line_final_price, subtotal=line_subtotal, vat_amount=line_vat
                            )

                            product.stock_quantity -= quantity
                            product.save(update_fields=['stock_quantity'])

                            total_subtotal += line_subtotal
                            total_vat += line_vat

                    if not items_added:
                         messages.error(request, "Πρέπει να προσθέσετε τουλάχιστον ένα προϊόν.")
                         raise Exception("No items added")

                    receipt.subtotal = total_subtotal
                    receipt.vat_amount = total_vat
                    receipt.total_amount = total_subtotal + total_vat
                    receipt.save()

                messages.success(request, f"Η απόδειξη {receipt.receipt_number} εκδόθηκε επιτυχώς.")
                return redirect('retail_receipt_detail', pk=receipt.pk) # Ανακατεύθυνση στη νέα απόδειξη

            except Exception as e:
                if str(e) not in ["Retail customer not found", "Insufficient stock", "No items added"]:
                     messages.error(request, f"Προέκυψε ένα άγνωστο σφάλμα: {e}")
        else:
            messages.error(request, "Η φόρμα περιέχει σφάλματα.")

        # Αν η φόρμα είναι άκυρη, ξαναστέλνουμε τις φόρμες στο template
        context = {
            'title': 'Νέα Πώληση Λιανικής (POS)',
            'form': form,
            'formset': formset
        }
        return render(request, 'core/retail_pos_form.html', context)

    else: # GET Request
        form = RetailReceiptForm()
        formset = RetailReceiptItemFormSet(prefix='items') # Προσθέτουμε το σωστό prefix

    context = {
        'title': 'Νέα Πώληση Λιανικής (POS)',
        'form': form,
        'formset': formset
    }
    return render(request, 'core/retail_pos_form.html', context)
@staff_member_required
def retail_receipt_list_view(request):
    receipts_qs = RetailReceipt.objects.select_related('customer').order_by('-issue_date', '-pk')
    
    paginator = Paginator(receipts_qs, 25) # 25 αποδείξεις ανά σελίδα
    page_number = request.GET.get('page')
    receipts_page = paginator.get_page(page_number)

    context = {
        'title': 'Λίστα Αποδείξεων Λιανικής',
        'receipts': receipts_page
    }
    return render(request, 'core/retail_receipt_list.html', context)


@staff_member_required
@require_POST
def retail_receipt_delete_view(request, pk):
    """
    Διαγράφει μια απόδειξη λιανικής και επαναφέρει το απόθεμα των προϊόντων.
    """
    receipt = get_object_or_404(RetailReceipt, pk=pk)
    receipt_number_for_message = receipt.receipt_number

    try:
        with transaction.atomic():
            # Κρίσιμο βήμα: Επαναφορά του αποθέματος για κάθε είδος της απόδειξης
            for item in receipt.items.all():
                if item.product:
                    # Χρησιμοποιούμε select_for_update για να "κλειδώσουμε" το προϊόν
                    # και να αποφύγουμε race conditions.
                    product_to_update = Product.objects.select_for_update().get(pk=item.product.pk)
                    product_to_update.stock_quantity += item.quantity
                    product_to_update.save(update_fields=['stock_quantity'])

            # Διαγραφή της απόδειξης
            receipt.delete()
            
            messages.success(request, f"Η απόδειξη {receipt_number_for_message} διαγράφηκε και το απόθεμα των προϊόντων επαναφέρθηκε.")
    
    except Product.DoesNotExist:
        messages.error(request, "Σφάλμα: Ένα από τα προϊόντα της απόδειξης δεν βρέθηκε και η διαγραφή απέτυχε.")
    except Exception as e:
        messages.error(request, f"Προέκυψε ένα μη αναμενόμενο σφάλμα: {e}")

    return redirect('retail_receipt_list')
@staff_member_required
def retail_receipt_detail_view(request, pk):
    receipt = get_object_or_404(RetailReceipt.objects.select_related('customer'), pk=pk)
    context = {
        'title': f'Απόδειξη Λιανικής: {receipt.receipt_number}',
        'receipt': receipt
    }
    return render(request, 'core/retail_receipt_detail.html', context)


@staff_member_required
def retail_receipt_pdf_view(request, pk):
    if not WEASYPRINT_AVAILABLE:
        raise Http404("Η βιβλιοθήκη WeasyPrint λείπει.")

    receipt = get_object_or_404(RetailReceipt.objects.select_related('customer'), pk=pk)
    company_info = getattr(settings, 'COMPANY_INFO', {})

    context = {
        'receipt': receipt,
        'company_info': company_info,
    }
    html_string = render_to_string('core/retail_receipt_pdf.html', context)
    html = HTML(string=html_string)
    
    response = HttpResponse(html.write_pdf(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="receipt_{receipt.receipt_number}.pdf"'
    return response
@login_required
def export_retail_receipts_to_excel(request):
    """
    Εξάγει μια λίστα με τις αποδείξεις λιανικής σε αρχείο Excel.
    """
    receipts_qs = RetailReceipt.objects.select_related('customer').all().order_by('-issue_date')

    data_for_export = []
    for receipt in receipts_qs:
        data_for_export.append({
            'Αρ. Απόδειξης': receipt.receipt_number,
            'Ημερομηνία & Ώρα': receipt.issue_date.strftime('%Y-%m-%d %H:%M:%S'),
            'Πελάτης': str(receipt.customer),
            'Καθαρή Αξία (€)': receipt.subtotal,
            'Ποσό ΦΠΑ (€)': receipt.vat_amount,
            'Τελικό Ποσό (€)': receipt.total_amount,
            'Κατάσταση': receipt.get_status_display(),
        })
    
    if not data_for_export:
        messages.warning(request, "Δεν βρέθηκαν αποδείξεις λιανικής για εξαγωγή.")
        return redirect(request.META.get('HTTP_REFERER', 'retail_receipt_list'))

    df = pd.DataFrame(data_for_export)
    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Αποδείξεις Λιανικής')
        worksheet = writer.sheets['Αποδείξεις Λιανικής']
        
        for idx, col in enumerate(df.columns):
            series = df[col]
            # --- ΕΔΩ ΕΙΝΑΙ Η ΔΙΟΡΘΩΣΗ ---
            # Διορθώθηκε η συντακτική δομή του υπολογισμού max_len
            max_len = max(
                series.astype(str).map(len).max(),
                len(str(series.name))
            ) + 2
            # --- ΤΕΛΟΣ ΔΙΟΡΘΩΣΗΣ ---
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len

    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="retail_receipts_export.xlsx"'
    return response
@login_required
def export_credit_notes_to_excel(request):
    """
    Εξάγει μια λίστα με τα πιστωτικά τιμολόγια σε αρχείο Excel.
    """
    credit_notes_qs = CreditNote.objects.select_related('customer', 'original_invoice').all().order_by('-issue_date')

    # Εδώ θα μπορούσαμε να προσθέσουμε τα φίλτρα από τη σελίδα της λίστας, αν υπήρχαν.
    # Προς το παρόν, εξάγει όλα τα πιστωτικά.

    data_for_export = []
    for cn in credit_notes_qs:
        data_for_export.append({
            'Αρ. Πιστωτικού': cn.credit_note_number,
            'Ημερομηνία Έκδοσης': cn.issue_date,
            'Πελάτης': str(cn.customer),
            'Αρχικό Τιμολόγιο': cn.original_invoice.invoice_number if cn.original_invoice else '-',
            'Αιτιολογία': cn.reason,
            'Καθαρή Αξία (€)': cn.subtotal,
            'Ποσό ΦΠΑ (€)': cn.vat_amount,
            'Συνολική Πίστωση (€)': cn.total_amount,
            'Κατάσταση': cn.get_status_display(),
        })
    
    if not data_for_export:
        messages.warning(request, "Δεν βρέθηκαν πιστωτικά τιμολόγια για εξαγωγή.")
        return redirect(request.META.get('HTTP_REFERER', 'credit_note_list'))

    df = pd.DataFrame(data_for_export)
    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Πιστωτικά Τιμολόγια')
        worksheet = writer.sheets['Πιστωτικά Τιμολόγια']
        
        for idx, col in enumerate(df.columns):
            series = df[col]
            if 'Ημερομηνία' in str(series.name):
                max_len = 20
            else:
                max_len = max(
                    series.astype(str).map(len).max(),
                    len(str(series.name))
                ) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len

    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="credit_notes_export.xlsx"'
    return response

@login_required
def delivery_note_list(request):
    notes_qs = DeliveryNote.objects.select_related('customer__parent', 'order', 'invoice').all()

    # Παίρνουμε ΟΛΕΣ τις τιμές των φίλτρων από το URL
    customer_filter_id = request.GET.get('customer_filter')
    status_filter = request.GET.get('status_filter')
    query_text = request.GET.get('q')
    
    # --- ΝΕΕΣ ΜΕΤΑΒΛΗΤΕΣ ΓΙΑ ΗΜΕΡΟΜΗΝΙΕΣ ---
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')

    # Εφαρμόζουμε τα φίλτρα ένα-ένα
    if query_text:
        notes_qs = notes_qs.filter(
            Q(delivery_note_number__icontains=query_text) | 
            Q(customer__company_name__icontains=query_text) |
            Q(customer__first_name__icontains=query_text) |
            Q(customer__last_name__icontains=query_text) |
            Q(shipping_name__icontains=query_text)
        ).distinct()
        
    if customer_filter_id:
        notes_qs = notes_qs.filter(
            Q(customer_id=customer_filter_id) | 
            Q(customer__parent_id=customer_filter_id)
        )
        
    if status_filter:
        notes_qs = notes_qs.filter(status=status_filter)
        
    # --- ΝΕΑ ΛΟΓΙΚΗ ΦΙΛΤΡΑΡΙΣΜΑΤΟΣ ΗΜΕΡΟΜΗΝΙΑΣ ---
    if date_from_str:
        try:
            notes_qs = notes_qs.filter(issue_date__gte=date_from_str)
        except: pass
        
    if date_to_str:
        try:
            notes_qs = notes_qs.filter(issue_date__lte=date_to_str)
        except: pass
    # --- ΤΕΛΟΣ ΝΕΑΣ ΛΟΓΙΚΗΣ ---

    # Ταξινομούμε τα τελικά αποτελέσματα
    notes_qs = notes_qs.order_by('-issue_date', '-pk')

    customers_for_filter = Customer.objects.filter(is_branch=False).order_by('company_name', 'last_name')
    delivery_note_statuses = DeliveryNote.Status.choices

    paginator = Paginator(notes_qs, 25)
    page_number = request.GET.get('page')
    notes = paginator.get_page(page_number)

    context = {
        'title': 'Λίστα Δελτίων Αποστολής',
        'notes': notes,
        'customers_for_filter': customers_for_filter,
        'delivery_note_statuses': delivery_note_statuses,
        'query_text': query_text,
        'customer_filter_id': int(customer_filter_id) if customer_filter_id else None,
        'status_filter': status_filter,
        # --- ΝΕΕΣ ΜΕΤΑΒΛΗΤΕΣ ΓΙΑ ΤΟ TEMPLATE ---
        'date_from_value': date_from_str,
        'date_to_value': date_to_str,
    }
    return render(request, 'core/delivery_note_list.html', context)


@login_required
def delivery_note_detail(request, pk):
    """
    Εμφανίζει τις λεπτομέρειες ενός συγκεκριμένου Δελτίου Αποστολής.
    """
    delivery_note = get_object_or_404(
        DeliveryNote.objects.select_related('customer', 'order'), 
        pk=pk
    )
    context = {
        'title': f'Δελτίο Αποστολής: {delivery_note.delivery_note_number}',
        'note': delivery_note,
    }
    return render(request, 'core/delivery_note_detail.html', context)


@login_required
def delivery_note_pdf_view(request, pk):
    """
    Δημιουργεί μια εκτυπώσιμη μορφή PDF για ένα Δελτίο Αποστολής.
    """
    if not WEASYPRINT_AVAILABLE:
        raise Http404("Η βιβλιοθήκη WeasyPrint λείπει. Η δημιουργία PDF δεν είναι δυνατή.")

    delivery_note = get_object_or_404(DeliveryNote.objects.select_related('customer__parent', 'order'), pk=pk)
   
    company_info = getattr(settings, 'COMPANY_INFO', {})

    context = {
        'note': delivery_note,
        'company_info': company_info,
    }

    html_string = render_to_string('core/delivery_note_pdf.html', context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))

    response = HttpResponse(html.write_pdf(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="DA_{delivery_note.delivery_note_number}.pdf"'
    return response
@login_required
def delivery_note_edit_view(request, pk):
    """
    Επεξεργασία ενός Δελτίου Αποστολής (κυρίως για αλλαγή κατάστασης).
    """
    delivery_note = get_object_or_404(DeliveryNote, pk=pk)

    if delivery_note.status == DeliveryNote.Status.CANCELLED:
        messages.error(request, "Δεν μπορείτε να επεξεργαστείτε ένα ακυρωμένο Δελτίο Αποστολής.")
        return redirect('delivery_note_detail', pk=delivery_note.pk)

    if request.method == 'POST':
        form = DeliveryNoteEditForm(request.POST, instance=delivery_note)
        if form.is_valid():
            form.save()
            messages.success(request, f"Οι αλλαγές στο Δελτίο Αποστολής {delivery_note.delivery_note_number} αποθηκεύτηκαν.")
            return redirect('delivery_note_detail', pk=delivery_note.pk)
    else:
        form = DeliveryNoteEditForm(instance=delivery_note)

    context = {
        'title': f'Επεξεργασία Δ.Α. {delivery_note.delivery_note_number}',
        'form': form,
        'note': delivery_note
    }
    return render(request, 'core/delivery_note_edit.html', context)
@login_required
def standalone_delivery_note_create_view(request):
    if request.method == 'POST':
        form = StandaloneDeliveryNoteForm(request.POST)
        formset = StandaloneDeliveryNoteItemFormSet(request.POST, prefix='items')

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                customer = form.cleaned_data.get('customer')
                if not customer:
                    customer, created = Customer.objects.get_or_create(
                        code='DN-GENERIC',
                        defaults={'first_name': 'Αυτόνομη', 'last_name': 'Διακίνηση'}
                    )
                
                dn = form.save(commit=False)
                dn.customer = customer
                dn.save()

                for item_form in formset.cleaned_data:
                    if item_form and (item_form.get('product') or item_form.get('description')):
                        product_instance = item_form.get('product')
                        desc = item_form.get('description') or (product_instance.name if product_instance else '')
                        
                        DeliveryNoteItem.objects.create(
                            delivery_note=dn,
                            product=product_instance,
                            quantity=item_form['quantity'],
                            description=desc
                        )
            
            messages.success(request, f"Το αυτόνομο Δελτίο Αποστολής {dn.delivery_note_number} δημιουργήθηκε.")
            return redirect('delivery_note_detail', pk=dn.pk)
        else:
            messages.error(request, "Η φόρμα περιέχει σφάλματα. Παρακαλώ διορθώστε τα.")
    else:
        # --- Η ΣΗΜΑΝΤΙΚΗ ΓΡΑΜΜΗ ΓΙΑ ΤΗΝ ΗΜΕΡΟΜΗΝΙΑ ---
        form = StandaloneDeliveryNoteForm(initial={'issue_date': timezone.now().date()})
        formset = StandaloneDeliveryNoteItemFormSet(prefix='items')

    context = {
        'title': 'Δημιουργία Αυτόνομου Δελτίου Αποστολής',
        'form': form,
        'formset': formset
    }
    return render(request, 'core/standalone_delivery_note_form.html', context)
@login_required
def invoice_edit_view(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    if request.method == 'POST':
        form = InvoiceEditForm(request.POST, instance=invoice)
        if form.is_valid():
            form.save()
            messages.success(request, f"Οι αλλαγές στο τιμολόγιο {invoice.invoice_number} αποθηκεύτηκαν.")
            return redirect('invoice_detail', pk=invoice.pk)
    else:
        form = InvoiceEditForm(instance=invoice)

    context = {
        'title': f'Επεξεργασία Τιμολογίου {invoice.invoice_number}',
        'form': form,
        'invoice': invoice
    }
    return render(request, 'core/invoice_edit.html', context)

@login_required
def view_invoice_cancellation_pdf(request, pk):
    """
    Δημιουργεί ένα εκτυπώσιμο PDF που επιβεβαιώνει την ακύρωση ενός τιμολογίου.
    """
    if not WEASYPRINT_AVAILABLE:
        raise Http404("Η βιβλιοθήκη WeasyPrint λείπει.")

    invoice = get_object_or_404(Invoice, pk=pk)

    # Επιτρέπουμε την εκτύπωση μόνο για ακυρωμένα τιμολόγια
    if invoice.status != Invoice.STATUS_CANCELLED:
        messages.error(request, "Μπορεί να εκτυπωθεί ακυρωτικό σημείωμα μόνο για ακυρωμένα τιμολόγια.")
        return redirect('invoice_detail', pk=invoice.pk)
        
    company_info = getattr(settings, 'COMPANY_INFO', {})

    context = {
        'invoice': invoice,
        'company_info': company_info,
    }

    html_string = render_to_string('core/invoice_cancellation_pdf.html', context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))

    response = HttpResponse(html.write_pdf(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="CANCELLATION_{invoice.invoice_number}.pdf"'
    return response
@login_required
def view_delivery_note_cancellation_pdf(request, pk):
    if not WEASYPRINT_AVAILABLE:
        raise Http404("Η βιβλιοθήκη WeasyPrint λείπει.")
    note = get_object_or_404(DeliveryNote.objects.select_related('customer__parent'), pk=pk)
    if note.status != DeliveryNote.Status.CANCELLED:
        messages.error(request, "Μπορεί να εκτυπωθεί ακυρωτικό σημείωμα μόνο για ακυρωμένα Δελτία Αποστολής.")
        return redirect('delivery_note_detail', pk=note.pk)
    company_info = getattr(settings, 'COMPANY_INFO', {})
    context = {'note': note, 'company_info': company_info}
    html_string = render_to_string('core/delivery_note_cancellation_pdf.html', context)
    html = HTML(string=html_string)
    response = HttpResponse(html.write_pdf(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="CANCELLATION_DA_{note.delivery_note_number}.pdf"'
    return response
@login_required
@require_POST
def create_invoice_from_delivery_note(request, dn_pk):
    dn = get_object_or_404(DeliveryNote.objects.select_related('customer__parent'), pk=dn_pk)

    # Έλεγχος 1: Είναι ΔΑ υποκαταστήματος;
    if not dn.customer.is_branch or not dn.customer.parent:
        messages.error(request, "Η τιμολόγηση από Δ.Α. επιτρέπεται μόνο για υποκαταστήματα με δηλωμένο κεντρικό.")
        return redirect('delivery_note_detail', pk=dn.pk)

    # Έλεγχος 2: Έχει ήδη τιμολογηθεί;
    if hasattr(dn, 'invoice') and dn.invoice:
        messages.warning(request, f"Αυτό το Δ.Α. έχει ήδη τιμολογηθεί (Τιμολόγιο: {dn.invoice.invoice_number}).")
        return redirect('invoice_detail', pk=dn.invoice.pk)

    parent_customer = dn.customer.parent

    with transaction.atomic():
        # Δημιουργούμε το τιμολόγιο στο όνομα του κεντρικού
        new_invoice = Invoice.objects.create(
            customer=parent_customer,
            delivery_note=dn, # Συνδέουμε το ΔΑ με το νέο τιμολόγιο
            issue_date=timezone.now().date(),
            status=Invoice.STATUS_ISSUED,
            purpose=dn.purpose,
            carrier=dn.carrier,
            license_plate=dn.license_plate
        )

        # Αντιγράφουμε τα είδη από το ΔΑ στο τιμολόγιο
        for item in dn.items.all():
            if item.product:
                InvoiceItem.objects.create(
                    invoice=new_invoice,
                    product=item.product,
                    description=item.description,
                    quantity=item.quantity,
                    unit_price=item.product.price, # Παίρνει την τρέχουσα τιμή
                    vat_percentage=item.product.vat_percentage,
                    total_price=item.quantity * item.product.price,
                    vat_amount=(item.quantity * item.product.price) * (item.product.vat_percentage / Decimal('100'))
                )

        # Υπολογίζουμε τα τελικά σύνολα και ενημερώνουμε το υπόλοιπο
        new_invoice.calculate_totals()
        new_invoice.save()
        parent_customer.balance += new_invoice.total_amount
        parent_customer.save(update_fields=['balance', 'updated_at'])

    messages.success(request, f"Δημιουργήθηκε το τιμολόγιο {new_invoice.invoice_number} από το Δ.Α. {dn.delivery_note_number}.")
    return redirect('invoice_detail', pk=new_invoice.pk)
@login_required
@require_POST
def create_delivery_note_from_order(request, order_pk):
    """
    Δημιουργεί ένα Δελτίο Αποστολής ΜΟΝΟ από μια 'Ολοκληρωμένη' παραγγελία.
    """
    order = get_object_or_404(Order, pk=order_pk)

    # Έλεγχος 1: Αν η παραγγελία ΔΕΝ είναι 'Ολοκληρωμένη', σταμάτα.
    if order.status != Order.STATUS_COMPLETED:
        messages.error(request, "Μπορούν να δημιουργηθούν Δελτία Αποστολής μόνο για 'Ολοκληρωμένες' παραγγελίες.")
        return redirect('order_detail', pk=order.pk)

    # Έλεγχος 2: Αν υπάρχει ήδη ενεργό Δ.Α., σταμάτα.
    if order.delivery_notes.filter(status__in=['PREPARING', 'SHIPPED', 'DELIVERED']).exists():
        existing_dn = order.delivery_notes.filter(status__in=['PREPARING', 'SHIPPED', 'DELIVERED']).first()
        messages.warning(request, f"Αυτή η παραγγελία έχει ήδη ενεργό Δελτίο Αποστολής (Αρ: {existing_dn.delivery_note_number}).")
        return redirect('delivery_note_detail', pk=existing_dn.pk)
        
    # Αν περάσουν οι έλεγχοι, προχωρά στη δημιουργία
    try:
        with transaction.atomic():
            delivery_note = DeliveryNote.objects.create(
                order=order,
                customer=order.customer,
                shipping_name=order.shipping_name,
                shipping_address=order.shipping_address,
                shipping_city=order.shipping_city,
                shipping_postal_code=order.shipping_postal_code,
                shipping_vat_number=order.customer.vat_number,
                purpose=order.purpose,
                carrier=order.carrier,
                license_plate=order.license_plate,
                issue_date=timezone.now().date(),
                notes=f"Βάσει Παραγγελίας {order.order_number}"
            )

            for item in order.items.all():
                DeliveryNoteItem.objects.create(
                    delivery_note=delivery_note,
                    product=item.product,
                    description=item.product.name if item.product else "Είδος χωρίς προϊόν",
                    quantity=item.quantity
                )
            
            messages.success(request, f"Το Δελτίο Αποστολής {delivery_note.delivery_note_number} δημιουργήθηκε επιτυχώς.")
            return redirect('delivery_note_detail', pk=delivery_note.pk)

    except Exception as e:
        messages.error(request, f"Προέκυψε ένα μη αναμενόμενο σφάλμα κατά τη δημιουργία του Δελτίου Αποστολής: {e}")
        return redirect('order_detail', pk=order.pk)
@login_required
@require_POST
def delivery_note_cancel_view(request, pk):
    """
    Ακυρώνει ένα Δελτίο Αποστολής. Αν το Δ.Α. προέρχεται από παραγγελία,
    ακυρώνει ΠΑΝΤΑ και τη σχετική παραγγελία για να επιστραφεί το απόθεμα.
    """
    note = get_object_or_404(DeliveryNote, pk=pk)
    
    if note.status in [DeliveryNote.Status.DELIVERED, DeliveryNote.Status.CANCELLED]:
        messages.error(request, f"Το Δελτίο Αποστολής είναι σε κατάσταση '{note.get_status_display()}' και δεν μπορεί να ακυρωθεί.")
        return redirect('delivery_note_detail', pk=note.pk)
    
    if hasattr(note, 'invoice') and note.invoice:
        messages.error(request, f"Το Δ.Α. {note.delivery_note_number} δεν μπορεί να ακυρωθεί γιατί έχει ήδη τιμολογηθεί.")
        return redirect('delivery_note_detail', pk=note.pk)

    try:
        with transaction.atomic():
            note.status = DeliveryNote.Status.CANCELLED
            note.save(update_fields=['status'])

            # --- ΤΕΛΙΚΗ ΚΑΙ ΓΕΝΙΚΗ ΛΟΓΙΚΗ ---
            # Αν το Δ.Α. έχει συνδεδεμένη παραγγελία, την ακυρώνουμε ΠΑΝΤΑ.
            if note.order:
                related_order = note.order
                related_order.status = Order.STATUS_CANCELLED
                related_order.save(update_fields=['status'])
                
                messages.info(request, f"Η σχετική παραγγελία {related_order.order_number} ακυρώθηκε και το απόθεμα ενημερώθηκε.")
            # --- ΤΕΛΟΣ ΛΟΓΙΚΗΣ ---

        messages.success(request, f"Το Δελτίο Αποστολής {note.delivery_note_number} ακυρώθηκε επιτυχώς.")

    except Exception as e:
        messages.error(request, f"Προέκυψε ένα σφάλμα: {e}")

    return redirect('delivery_note_detail', pk=note.pk)            
                            
@login_required
def reporting_hub_view(request):
    """
    Εμφανίζει την κεντρική σελίδα με τις διαθέσιμες αναφορές.
    """
    return render(request, 'core/reporting_hub.html', {'title': 'Κέντρο Αναφορών'})

@login_required
def report_sales_by_month_view(request):
    # Δημιουργούμε μια λίστα με τους τελευταίους 12 μήνες, με αρχική αξία 0
    sales_by_month = {}
    today = timezone.now().date()
    for i in range(12):
        # Ξεκινάμε από τον τρέχοντα μήνα και πάμε 11 μήνες πίσω
        month_date = today - relativedelta(months=i)
        # Το κλειδί είναι η αρχή του μήνα (π.χ. '2025-06-01')
        sales_by_month[month_date.replace(day=1)] = Decimal('0.00')

    # Παίρνουμε τα πραγματικά δεδομένα από τη βάση
    sales_data_from_db = Invoice.objects.filter(
        status__in=[Invoice.STATUS_ISSUED, Invoice.STATUS_PAID],
        issue_date__gte=list(sales_by_month.keys())[-1] # Από τον παλαιότερο μήνα και μετά
    ).annotate(
        month=TruncMonth('issue_date')
    ).values('month').annotate(total_sales=Sum('total_amount')).order_by('month')

    # Ενημερώνουμε το λεξικό μας με τις πραγματικές πωλήσεις
    for entry in sales_data_from_db:
        if entry['month'] in sales_by_month:
            sales_by_month[entry['month']] = entry['total_sales']

    # Ταξινομούμε τους μήνες χρονολογικά για το γράφημα και τον πίνακα
    sorted_months = sorted(sales_by_month.items())

    # Προετοιμάζουμε τα δεδομένα για τη Chart.js
    labels = [date.strftime('%b %Y') for date, sales in sorted_months]
    data_points = [float(sales) for date, sales in sorted_months]

    # Προετοιμάζουμε τα δεδομένα και για τον πίνακα στο template
    table_data = [{'month': date, 'sales': sales} for date, sales in reversed(sorted_months)]

    context = {
        'title': 'Αναφορά Πωλήσεων ανά Μήνα',
        'labels': json.dumps(labels),
        'data': json.dumps(data_points),
        'table_data': table_data, # Νέα προσθήκη για τον πίνακα
    }
    return render(request, 'core/report_sales_by_month.html', context)
@login_required
def report_customer_balance_view(request):
    today = timezone.now().date()
    default_from = today.replace(day=1)
    
    date_from_str = request.GET.get('date_from', default_from.strftime('%Y-%m-%d'))
    date_to_str = request.GET.get('date_to', today.strftime('%Y-%m-%d'))
    
    # --- ΑΛΛΑΓΗ: Παίρνουμε customer_id αντί για q ---
    customer_id = request.GET.get('customer')
    sales_rep_id = request.GET.get('sales_rep')

    customers_qs = Customer.objects.filter(is_branch=False)

    if sales_rep_id:
        customers_qs = customers_qs.filter(sales_rep_id=sales_rep_id)
    
    # --- ΑΛΛΑΓΗ: Φιλτράρουμε με βάση το ID του πελάτη ---
    if customer_id:
        customers_qs = customers_qs.filter(pk=customer_id)

    report_data = []
    # (Η υπόλοιπη λογική υπολογισμού παραμένει ακριβώς ίδια)
    for customer in customers_qs:
        debits_before = customer.invoices.filter(status__in=[Invoice.STATUS_ISSUED, Invoice.STATUS_PAID], issue_date__lt=date_from_str).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        payments_before = customer.payments.filter(status=Payment.STATUS_ACTIVE, payment_date__lt=date_from_str).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
        credits_before = customer.credit_notes.filter(status=CreditNote.Status.ISSUED, issue_date__lt=date_from_str).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        balance_brought_forward = debits_before - (payments_before + credits_before)
        debits_in_period = customer.invoices.filter(status__in=[Invoice.STATUS_ISSUED, Invoice.STATUS_PAID], issue_date__range=[date_from_str, date_to_str]).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        payments_in_period = customer.payments.filter(status=Payment.STATUS_ACTIVE, payment_date__range=[date_from_str, date_to_str]).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
        credits_in_period = customer.credit_notes.filter(status=CreditNote.Status.ISSUED, issue_date__range=[date_from_str, date_to_str]).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        credit_in_period = payments_in_period + credits_in_period
        closing_balance = balance_brought_forward + debits_in_period - credit_in_period
        
        # Τώρα δεν χρειάζεται ο έλεγχος, αφού θα φέρνουμε πάντα αποτελέσματα (εκτός αν δεν επιλεγεί πελάτης)
        report_data.append({
            'customer': customer, 'opening_balance': balance_brought_forward,
            'debit_in_period': debits_in_period, 'credit_in_period': credit_in_period,
            'closing_balance': closing_balance
        })

    report_data.sort(key=lambda x: x['closing_balance'], reverse=True)
    
    all_sales_reps = SalesRepresentative.objects.select_related('user').all()
    # --- ΠΡΟΣΘΗΚΗ: Παίρνουμε όλους τους πελάτες για το Select2 ---
    all_customers = Customer.objects.filter(is_branch=False).order_by('company_name', 'last_name')

    context = {
        'title': 'Αναφορά - Κίνηση Πελατών ανά Περίοδο',
        'report_data': report_data,
        'date_from_value': date_from_str,
        'date_to_value': date_to_str,
        'all_sales_reps': all_sales_reps,
        'sales_rep_id_value': int(sales_rep_id) if sales_rep_id else None,
        'all_customers': all_customers, # <-- Προσθήκη για το Select2
        'customer_id_value': int(customer_id) if customer_id else None,
    }
    return render(request, 'core/report_customer_balance.html', context)
@login_required
def report_vat_analysis_view(request):
    """
    Δημιουργεί την αναφορά ανάλυσης ΦΠΑ για μια δεδομένη χρονική περίοδο.
    """
    today = timezone.now().date()
    # Προεπιλεγμένη περίοδος: ο τρέχων μήνας
    date_from_str = request.GET.get('date_from', today.replace(day=1).strftime('%Y-%m-%d'))
    date_to_str = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    # 1. Υπολογισμός του ΦΠΑ εκροών (από Τιμολόγια)
    output_vat_qs = InvoiceItem.objects.filter(
        invoice__status__in=[Invoice.STATUS_ISSUED, Invoice.STATUS_PAID],
        invoice__issue_date__range=[date_from_str, date_to_str]
    )
    output_summary = output_vat_qs.values('vat_percentage').annotate(
        total_net=Sum('total_price'), 
        total_vat=Sum('vat_amount')
    ).order_by('vat_percentage')

    # 2. Υπολογισμός του ΦΠΑ εισροών (μείωση από Πιστωτικά)
    input_vat_qs = CreditNoteItem.objects.filter(
        credit_note__status=CreditNote.Status.ISSUED,
        credit_note__issue_date__range=[date_from_str, date_to_str]
    )
    input_summary = input_vat_qs.values('vat_percentage').annotate(
        total_net=Sum('total_price'), 
        total_vat=Sum('vat_amount')
    ).order_by('vat_percentage')

    # 3. Συνδυασμός των αποτελεσμάτων σε ένα λεξικό
    final_analysis = defaultdict(lambda: {'net': Decimal(0), 'vat': Decimal(0)})
    
    for item in output_summary:
        rate = item['vat_percentage']
        final_analysis[rate]['net'] += item['total_net']
        final_analysis[rate]['vat'] += item['total_vat']

    for item in input_summary:
        rate = item['vat_percentage']
        final_analysis[rate]['net'] -= item['total_net'] # Αφαιρούμε την αξία
        final_analysis[rate]['vat'] -= item['total_vat'] # Αφαιρούμε τον ΦΠΑ

    # Μετατροπή σε λίστα και υπολογισμός συνόλων
    analysis_list = [{'rate': rate, 'data': data} for rate, data in sorted(final_analysis.items())]
    
    total_net_value = sum(item['data']['net'] for item in analysis_list)
    total_vat_value = sum(item['data']['vat'] for item in analysis_list)

    context = {
        'title': 'Αναφορά - Ανάλυση ΦΠΑ',
        'analysis_data': analysis_list,
        'total_net_value': total_net_value,
        'total_vat_value': total_vat_value,
        'date_from_value': date_from_str,
        'date_to_value': date_to_str,
    }
    return render(request, 'core/report_vat_analysis.html', context)


@login_required
def report_profitability_view(request):
    today = timezone.now().date()
    default_from = today.replace(day=1)
    
    date_from_str = request.GET.get('date_from', default_from.strftime('%Y-%m-%d'))
    date_to_str = request.GET.get('date_to', today.strftime('%Y-%m-%d'))
    
    # --- ΝΕΑ ΠΡΟΣΘΗΚΗ: Παίρνουμε το ID του προϊόντος από το φίλτρο ---
    product_id = request.GET.get('product')

    # Το βασικό queryset
    items_qs = InvoiceItem.objects.filter(
        invoice__status__in=[Invoice.STATUS_ISSUED, Invoice.STATUS_PAID],
        invoice__issue_date__range=[date_from_str, date_to_str],
        product__cost_price__isnull=False,
        product__cost_price__gt=0
    )

    # --- ΝΕΑ ΠΡΟΣΘΗΚΗ: Εφαρμόζουμε το φίλτρο του προϊόντος αν υπάρχει ---
    if product_id:
        items_qs = items_qs.filter(product_id=product_id)

    # Η υπόλοιπη λογική παραμένει ως έχει
    profit_data = items_qs.values(
        'product__name', 'product__code'
    ).annotate(
        total_quantity=Coalesce(Sum('quantity'), Decimal('0.0')),
        total_revenue=Coalesce(Sum(F('quantity') * F('unit_price') * (Decimal(1) - F('discount_percentage') / 100)), Decimal('0.0')),
        total_cost=Coalesce(Sum(F('quantity') * F('product__cost_price')), Decimal('0.0'))
    ).annotate(
        total_profit=F('total_revenue') - F('total_cost')
    ).annotate(
        profit_margin=Case(
            When(total_revenue__gt=0, 
                 then=ExpressionWrapper((F('total_revenue') - F('total_cost')) * 100 / F('total_revenue'), output_field=DecimalField())),
            default=Value(Decimal('0.0')),
            output_field=DecimalField()
        )
    ).order_by('-total_profit')

    # --- ΝΕΑ ΠΡΟΣΘΗΚΗ: Παίρνουμε όλα τα προϊόντα για το dropdown ---
    all_products = Product.objects.filter(is_active=True).order_by('name')

    context = {
        'title': 'Αναφορά - Ανάλυση Κερδοφορίας ανά Προϊόν',
        'profit_data': profit_data,
        'date_from_value': date_from_str,
        'date_to_value': date_to_str,
        'all_products': all_products, # <-- Προσθήκη για το Select2
        'product_id_value': int(product_id) if product_id else None,
    }
    return render(request, 'core/report_profitability.html', context)


@login_required
def report_sales_by_rep_view(request):
    today = timezone.now().date()
    default_from = today.replace(day=1)
    
    date_from_str = request.GET.get('date_from', default_from.strftime('%Y-%m-%d'))
    date_to_str = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    # --- ΝΕΑ ΛΟΓΙΚΗ: ΞΕΚΙΝΑΜΕ ΑΠΟ ΤΙΣ ΠΡΟΜΗΘΕΙΕΣ ---
    
    # Το βασικό μας query ξεκινάει από το μοντέλο Commission
    sales_data = Commission.objects.filter(
        # Φιλτράρουμε τις προμήθειες με βάση την ημερομηνία του Τιμολογίου τους
        invoice__issue_date__range=[date_from_str, date_to_str]
    ).values(
        # Ομαδοποιούμε ανά πωλητή
        'sales_rep__user__first_name',
        'sales_rep__user__last_name'
    ).annotate(
        # Αθροίζουμε τα πεδία που θέλουμε
        invoice_count=Count('invoice__id', distinct=True), # Μετράμε τα μοναδικά τιμολόγια
        total_net_sales=Sum('invoice__subtotal'),      # Παίρνουμε το τζίρο από το σχετικό τιμολόγιο
        total_commission=Sum('calculated_amount')      # Παίρνουμε την προμήθεια από την ίδια την εγγραφή
    ).order_by('-total_net_sales')
    # --- ΤΕΛΟΣ ΝΕΑΣ ΛΟΓΙΚΗΣ ---

    context = {
        'title': 'Αναφορά - Πωλήσεις ανά Πωλητή',
        'sales_data': sales_data,
        'date_from_value': date_from_str,
        'date_to_value': date_to_str,
    }
    return render(request, 'core/report_sales_by_rep.html', context)
@login_required
def report_sales_by_city_view(request):
    today = timezone.now().date()
    default_from = today.replace(day=1)
    
    date_from_str = request.GET.get('date_from', default_from.strftime('%Y-%m-%d'))
    date_to_str = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    # Ομαδοποιούμε τα τιμολόγια ανά πόλη του πελάτη και αθροίζουμε την καθαρή αξία
    sales_by_city = Invoice.objects.filter(
        status__in=[Invoice.STATUS_ISSUED, Invoice.STATUS_PAID],
        issue_date__range=[date_from_str, date_to_str]
    ).values(
        'customer__city' # Ομαδοποίηση βάσει της πόλης του πελάτη
    ).annotate(
        total_net_sales=Sum('subtotal')
    ).order_by('-total_net_sales')

    context = {
        'title': 'Αναφορά - Πωλήσεις ανά Πόλη',
        'sales_data': sales_by_city,
        'date_from_value': date_from_str,
        'date_to_value': date_to_str,
    }
    return render(request, 'core/report_sales_by_city.html', context)
@login_required
def order_create_from_invoice_view(request, pk):
    """
    Δημιουργεί μια νέα παραγγελία, αντιγράφοντας τα στοιχεία από ένα υπάρχον τιμολόγιο.
    """
    original_invoice = get_object_or_404(Invoice.objects.select_related('customer'), pk=pk)

    # Δημιουργούμε ένα νέο, προσωρινό αντικείμενο παραγγελίας
    new_order = Order(
        customer=original_invoice.customer,
        shipping_name=original_invoice.shipping_name,
        shipping_address=original_invoice.shipping_address,
        shipping_city=original_invoice.shipping_city,
        shipping_postal_code=original_invoice.shipping_postal_code,
        # Ορίζουμε νέες τιμές για την ημερομηνία και την κατάσταση
        order_date=timezone.now().date(),
        status=Order.STATUS_PENDING 
    )

    # Προετοιμάζουμε τα είδη από το τιμολόγιο για το formset
    initial_items_data = []
    for item in original_invoice.items.all():
        initial_items_data.append({
            'product': item.product,
            'quantity': item.quantity,
            'is_gift': item.is_gift,
            'unit_price': item.unit_price,
            'discount_percentage': item.discount_percentage,
            'vat_percentage': item.vat_percentage,
        })

    form = OrderForm(instance=new_order)
    formset = OrderItemFormSet(initial=initial_items_data, prefix='items')

    messages.info(request, f"Αντιγραφή από το τιμολόγιο {original_invoice.invoice_number}. Ελέγξτε και αποθηκεύστε.")
    
    context = {
        'form': form,
        'formset': formset,
        'title': 'Δημιουργία Παραγγελίας από Τιμολόγιο',
    }
    return render(request, 'core/order_create.html', context)
@login_required
def order_copy_view(request, pk):
    """
    Δημιουργεί ένα αντίγραφο μιας υπάρχουσας παραγγελίας και προ-συμπληρώνει τη φόρμα δημιουργίας.
    """
    original_order = get_object_or_404(Order, pk=pk)
    
    # Δημιουργούμε ένα νέο αντικείμενο παραγγελίας στη μνήμη (δεν το σώζουμε)
    # αντιγράφοντας τα βασικά πεδία από την αρχική.
    new_order = Order(
        customer=original_order.customer,
        shipping_name=original_order.shipping_name,
        shipping_address=original_order.shipping_address,
        shipping_city=original_order.shipping_city,
        shipping_postal_code=original_order.shipping_postal_code,
        purpose=original_order.purpose,
        carrier=original_order.carrier,
        license_plate=original_order.license_plate,
        comments=original_order.comments,
        # Ορίζουμε νέες τιμές για την ημερομηνία και την κατάσταση
        order_date=timezone.now().date(),
        status=Order.STATUS_PENDING 
    )

    # Προετοιμάζουμε τα αρχικά δεδομένα για τα είδη (formset)
    initial_items_data = []
    for item in original_order.items.all():
        initial_items_data.append({
            'product': item.product,
            'quantity': item.quantity,
            'is_gift': item.is_gift,
            'unit_price': item.unit_price,
            'discount_percentage': item.discount_percentage,
            'vat_percentage': item.vat_percentage,
            'comments': item.comments,
        })

    # Δημιουργούμε τις φόρμες με τα προσυμπληρωμένα δεδομένα
    # Χρησιμοποιούμε το instance για τη βασική φόρμα και το initial για το formset
    form = OrderForm(instance=new_order)
    formset = OrderItemFormSet(initial=initial_items_data, prefix='items')

    messages.info(request, f"Αντίγραφο της παραγγελίας {original_order.order_number}. Ελέγξτε και αποθηκεύστε.")
    
    context = {
        'form': form,
        'formset': formset,
        'title': 'Δημιουργία Αντιγράφου Παραγγελίας',
    }
    # Χρησιμοποιούμε το template της δημιουργίας παραγγελίας για να εμφανίσουμε τις φόρμες
    return render(request, 'core/order_create.html', context)
@login_required
@require_POST
def ajax_change_order_status(request, pk):
    """
    View που δέχεται AJAX request για να αλλάξει την κατάσταση μιας παραγγελίας.
    """
    order = get_object_or_404(Order, pk=pk)
    new_status = request.POST.get('new_status')

    # Λίστα με τις επιτρεπτές καταστάσεις για ασφάλεια
    valid_statuses = [choice[0] for choice in Order.ORDER_STATUS_CHOICES]
    if new_status not in valid_statuses:
        return JsonResponse({'success': False, 'error': 'Μη έγκυρη κατάσταση.'}, status=400)

    # Έλεγχος αν η παραγγελία είναι κλειδωμένη (π.χ. έχει τιμολογηθεί)
    if hasattr(order, 'invoice') and order.invoice:
        return JsonResponse({'success': False, 'error': 'Η παραγγελία έχει τιμολογηθεί και δεν μπορεί να αλλάξει.'}, status=403)
        
    order.status = new_status
    order.save(update_fields=['status'])
    
    # Επιστρέφουμε την επιτυχία και το νέο display name της κατάστασης
    return JsonResponse({
        'success': True, 
        'new_status_display': order.get_status_display(),
        'new_status_code': order.status
        })
@login_required
@require_POST
def ajax_quick_stock_entry(request, pk):
    """
    Δέχεται AJAX request για γρήγορη καταχώρηση παραλαβής αποθέματος.
    """
    product = get_object_or_404(Product, pk=pk)
    quantity_str = request.POST.get('quantity_added')

    if not quantity_str:
        return JsonResponse({'success': False, 'error': 'Η ποσότητα είναι υποχρεωτική.'}, status=400)
    
    try:
        quantity = Decimal(quantity_str)
        if quantity <= 0:
            raise ValueError("Η ποσότητα πρέπει να είναι θετικός αριθμός.")
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': 'Παρακαλώ εισάγετε έναν έγκυρο αριθμό.'}, status=400)

    # Η δημιουργία του StockReceipt θα ενεργοποιήσει το signal που ενημερώνει το απόθεμα
    StockReceipt.objects.create(
        product=product,
        quantity_added=quantity,
        user_who_recorded=request.user,
        notes=f"Γρήγορη καταχώρηση από τη λίστα προϊόντων."
    )
    
    # Παίρνουμε την ενημερωμένη ποσότητα για να την επιστρέψουμε στο frontend
    product.refresh_from_db()

    return JsonResponse({
        'success': True,
        'new_stock': product.stock_quantity,
        'product_name': product.name
    })
@login_required
@require_POST
def ajax_send_invoice_email(request, pk):
    invoice = get_object_or_404(Invoice.objects.select_related('customer'), pk=pk)

    to_email = request.POST.get('to_email')
    subject = request.POST.get('subject')
    body = request.POST.get('body')

    if not to_email or not subject or not body:
        return JsonResponse({'success': False, 'error': 'Όλα τα πεδία είναι υποχρεωτικά.'}, status=400)

    try:
        # Δημιουργούμε το PDF στη μνήμη
        company_info = getattr(settings, 'COMPANY_INFO', {})
        html_string = render_to_string('core/invoice_pdf.html', {'invoice': invoice, 'company_info': company_info})
        pdf_file = HTML(string=html_string).write_pdf()

        # Δημιουργούμε το email
        email = EmailMessage(
            subject,
            body,
            settings.COMPANY_INFO.get('EMAIL'), # Από ποιον
            [to_email] # Προς ποιον
        )

        # Επισυνάπτουμε το PDF
        email.attach(f'invoice_{invoice.invoice_number}.pdf', pdf_file, 'application/pdf')

        # Στέλνουμε το email (θα τυπωθεί στην κονσόλα)
        email.send()

        return JsonResponse({'success': True, 'message': 'Το email στάλθηκε (εμφανίστηκε στην κονσόλα).'})

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Σφάλμα κατά την αποστολή: {str(e)}'}, status=500)
@login_required
@require_GET
def ajax_product_history_view(request, pk):
    product = get_object_or_404(Product, pk=pk)

    history_log = []

    # Πωλήσεις (από τιμολογημένα είδη)
    sales = InvoiceItem.objects.filter(product=product, invoice__status__in=[Invoice.STATUS_ISSUED, Invoice.STATUS_PAID])
    for item in sales:
        history_log.append({
            'date': item.invoice.issue_date,
            'type': 'Πώληση',
            'quantity': -item.quantity, # Αρνητικό γιατί μειώνει το απόθεμα
            'document': item.invoice,
            'document_number': item.invoice.invoice_number,
            'url_name': 'invoice_detail'
        })

    # Επιστροφές (από πιστωτικά)
    returns = CreditNoteItem.objects.filter(product=product, credit_note__status=CreditNote.Status.ISSUED)
    for item in returns:
        history_log.append({
            'date': item.credit_note.issue_date,
            'type': 'Επιστροφή',
            'quantity': item.quantity, # Θετικό γιατί αυξάνει το απόθεμα
            'document': item.credit_note,
            'document_number': item.credit_note.credit_note_number,
            'url_name': 'credit_note_detail'
        })

    # Παραλαβές
    receipts = StockReceipt.objects.filter(product=product)
    for item in receipts:
        history_log.append({
            'date': item.date_received.date(),
            'type': 'Παραλαβή',
            'quantity': item.quantity_added,
            'document': item,
            'document_number': f"Παραλαβή #{item.pk}",
            'url_name': 'stock_receipt_list' # Δεν υπάρχει detail view, οπότε πάμε στη λίστα
        })

    # Ταξινόμηση όλων των κινήσεων χρονολογικά
    history_log.sort(key=lambda x: x['date'], reverse=True)

    context = {'history_log': history_log, 'product': product}
    html = render_to_string('core/partials/_product_history_table.html', context)
    return JsonResponse({'html': html})
@login_required
def supplier_list_view(request):
    suppliers = Supplier.objects.all()
    title = 'Λίστα Προμηθευτών'
    context = {
        'title': title,
        'suppliers': suppliers
    }
    return render(request, 'core/supplier_list.html', context)
@login_required
def supplier_create_view(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save()
            messages.success(request, f'Ο προμηθευτής "{supplier.name}" δημιουργήθηκε επιτυχώς.')
            return redirect('supplier_list') # Επιστροφή στη λίστα μετά την επιτυχία
    else:
        form = SupplierForm()

    context = {
        'title': 'Δημιουργία Νέου Προμηθευτή',
        'form': form
    }
    return render(request, 'core/supplier_form.html', context)
@login_required
def supplier_detail_view(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    context = {
        'title': f'Καρτέλα Προμηθευτή: {supplier.name}',
        'supplier': supplier
    }
    return render(request, 'core/supplier_detail.html', context)

@login_required
def supplier_edit_view(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, 'Οι αλλαγές στον προμηθευτή αποθηκεύτηκαν επιτυχώς.')
            return redirect('supplier_detail', pk=supplier.pk)
    else:
        form = SupplierForm(instance=supplier)
    
    context = {
        'title': f'Επεξεργασία Προμηθευτή: {supplier.name}',
        'form': form,
    }
    # Επαναχρησιμοποιούμε το ίδιο template με τη δημιουργία
    return render(request, 'core/supplier_form.html', context)

@login_required
@require_POST # Ασφάλεια: επιτρέπουμε μόνο POST requests για διαγραφή
def supplier_delete_view(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    supplier_name = supplier.name
    # Εδώ θα μπορούσαμε να προσθέσουμε έλεγχο αν ο προμηθευτής έχει συνδεδεμένες Εντολές Αγοράς
    # και να απαγορεύσουμε τη διαγραφή. Το αφήνουμε για αργότερα.
    supplier.delete()
    messages.success(request, f'Ο προμηθευτής "{supplier_name}" διαγράφηκε επιτυχώς.')
    return redirect('supplier_list')
# --- VIEWS ΓΙΑ ΕΝΤΟΛΕΣ ΑΓΟΡΑΣ ---
@login_required
def purchase_order_list_view(request):
    purchase_orders = PurchaseOrder.objects.select_related('supplier').all()
    title = 'Λίστα Εντολών Αγοράς'
    
    # Θα προσθέσουμε φίλτρα εδώ αργότερα
    
    context = {
        'title': title,
        'purchase_orders': purchase_orders
    }
    return render(request, 'core/purchase_order_list.html', context)
 
@login_required
def purchase_order_create_view(request):
    # --- ΝΕΑ ΛΟΓΙΚΗ ΓΙΑ ΠΡΟΣΥΜΠΛΗΡΩΣΗ ---
    supplier_id = request.GET.get('supplier_id')
    initial_data = {}
    if supplier_id:
        initial_data['supplier'] = supplier_id
    # --- ΤΕΛΟΣ ΝΕΑΣ ΛΟΓΙΚΗΣ ---

    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST)
        formset = PurchaseOrderItemFormSet(request.POST, prefix='items')
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                po = form.save(commit=False)
                po.status = 'DRAFT' # Αλλάζουμε την αρχική κατάσταση σε Πρόχειρη
                po.save()
                
                formset.instance = po
                formset.save()
                
                total = po.items.aggregate(total_sum=Sum(F('quantity') * F('cost_price')))['total_sum'] or Decimal('0.00')
                po.total_amount = total
                po.save(update_fields=['total_amount'])

            messages.success(request, f'Η Εντολή Αγοράς {po.po_number} δημιουργήθηκε επιτυχώς.')
            return redirect('purchase_order_detail', pk=po.pk) # Ανακατεύθυνση στο detail
    else:
        # Χρησιμοποιούμε τα initial_data που φτιάξαμε
        form = PurchaseOrderForm(initial=initial_data)
        formset = PurchaseOrderItemFormSet(prefix='items', queryset=PurchaseOrderItem.objects.none())

    context = {
        'title': 'Δημιουργία Νέας Εντολής Αγοράς',
        'form': form,
        'formset': formset
    }
    return render(request, 'core/purchase_order_form.html', context)
@login_required
def purchase_order_detail_view(request, pk):
    po = get_object_or_404(PurchaseOrder.objects.select_related('supplier'), pk=pk)
    context = {
        'title': f'Εντολή Αγοράς: {po.po_number}',
        'po': po
    }
    return render(request, 'core/purchase_order_detail.html', context)                                   
login_required
def purchase_order_detail_view(request, pk):
    po = get_object_or_404(PurchaseOrder.objects.select_related('supplier'), pk=pk)
    context = {
        'title': f'Εντολή Αγοράς: {po.po_number}',
        'po': po
    }
    return render(request, 'core/purchase_order_detail.html', context)

@login_required
def purchase_order_edit_view(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST, instance=po)
        formset = PurchaseOrderItemFormSet(request.POST, instance=po, prefix='items')
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                po = form.save()
                formset.save()
                total = po.items.aggregate(total_sum=Sum('total_cost'))['total_sum'] or Decimal('0.00')
                po.total_amount = total
                po.save(update_fields=['total_amount'])
            messages.success(request, 'Η Εντολή Αγοράς ενημερώθηκε επιτυχώς.')
            return redirect('purchase_order_detail', pk=po.pk)
    else:
        form = PurchaseOrderForm(instance=po)
        formset = PurchaseOrderItemFormSet(instance=po, prefix='items')
    
    context = {
        'title': f'Επεξεργασία Εντολής Αγοράς: {po.po_number}',
        'form': form,
        'formset': formset,
        'po': po
    }
    return render(request, 'core/purchase_order_form.html', context)


@login_required
@require_POST
def purchase_order_delete_view(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    po_number = po.po_number
    # Προσοχή: Εδώ δεν υπάρχει λογική επαναφοράς αποθέματος,
    # καθώς υποθέτουμε ότι η διαγραφή γίνεται πριν την παραλαβή.
    po.delete()
    messages.success(request, f'Η Εντολή Αγοράς {po_number} διαγράφηκε επιτυχώς.')
    return redirect('purchase_order_list')
@login_required
def purchase_order_pdf_view(request, pk):
    if not WEASYPRINT_AVAILABLE:
        raise Http404("Η βιβλιοθήκη WeasyPrint λείπει. Η δημιουργία PDF δεν είναι δυνατή.")

    po = get_object_or_404(PurchaseOrder.objects.select_related('supplier'), pk=pk)
    company_info = getattr(settings, 'COMPANY_INFO', {})
    
    context = {
        'po': po,
        'company_info': company_info,
    }

    html_string = render_to_string('core/purchase_order_pdf.html', context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    
    response = HttpResponse(html.write_pdf(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="PO_{po.po_number}.pdf"'
    return response
@login_required
def export_purchase_orders_excel(request):
    # Εδώ θα μπορούσαμε να εφαρμόσουμε τα φίλτρα της λίστας αν υπήρχαν
    pos = PurchaseOrder.objects.select_related('supplier').all().order_by('-order_date')

    data_for_export = []
    for po in pos:
        data_for_export.append({
            'Αρ. Εντολής': po.po_number,
            'Ημερομηνία': po.order_date,
            'Προμηθευτής': po.supplier.name,
            'Κατάσταση': po.get_status_display(),
            'Συνολικό Ποσό (€)': po.total_amount,
        })
    
    if not data_for_export:
        messages.warning(request, "Δεν βρέθηκαν εντολές αγοράς για εξαγωγή.")
        return redirect('purchase_order_list')

    df = pd.DataFrame(data_for_export)
    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Εντολές Αγοράς')
        worksheet = writer.sheets['Εντολές Αγοράς']
        for idx, col in enumerate(df):
            series = df[col]
            max_len = max((series.astype(str).map(len).max(), len(str(series.name)))) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len
            
    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="purchase_orders_export.xlsx"'
    return response        